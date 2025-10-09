from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import get_column_letter
import npttf2utf
from openpyxl.utils import range_boundaries
import openpyxl
from openpyxl.styles import Alignment, Font


# File path and sheet name
file_path = r"C:\Users\Acer\Desktop\Software_Input.xlsx"
sheet_name = 'Analysis'
Kalimatifont_isset = True
Preetifont_isset = False


vastavikDarrate = "वास्तविक दररेट"
search_srtings = [
        "वास्तविक दररेट","वास्तविक दररेट",
        "१५% ठेकेदारको ओभरहेड", "१५% ठेकेदारको ओभरहेड", "१५% ठेकेदारको ओभरहेड", "१५% ठेकेदार ओभरहेड", "१५% ठेकेदार ओभरहेड","१५% ठेकेदार ओभरहेड ",
        "जम्मा दर रेट", "जम्मा दर रेट"
    ]
shramik = ["श्रमिक", "श्रमिक "]
nirmaSamagri = ["निर्माण सामग्री", "स्रोत साधन", "निर्माण सामग्री लगाउने ज्याला समेत"]
yantrikUpakaran = ["यान्त्रिक उपकरण", "यान्त्रिक उपकरण"]
rupees = "रु."
paisa = "पै."

if Preetifont_isset:
    vastavikDarrate = "jf:tljs b//]6"
    search_srtings = [ "jf:tljs b//]6",
                       "!%Ü 7]s]bf/sf] cf]e/x]8", "!%Ü 7]s]bf/sf] cf]e/x]8 ", "!%Ü 7]s]bf/ cf]e/x]8", "!%Ü 7]s]bf/ cf]e/x]8 "
                       "hDdf b/ /]6",
                       "jf:tljs b//]6",
                       "!%Ü 7]s]bf/sf] cf]e/x]8", "!%Ü 7]s]bf/sf] cf]e/x]8 ", "!%Ü 7]s]bf/ cf]e/x]8",
                       "hDdf b/ /]6"
                       ]
    shramik = [">lds", ">lds "]
    nirmaSamagri = ["lgdf{0f ;fdu|L", ";|f]t ;fwg", "निर्माण सामग्री लगाउने ज्याला समेत"]
    yantrikUpakaran = ["oflGqs pks/0f", "यान्त्रिक उपकरण"]
    rupees = "?="
    paisa = "k}="




# Load workbook
wb = load_workbook(file_path, data_only=True)
sheet = wb[sheet_name]

# Define exact column range: B to H
start_col = column_index_from_string('B')  # 2
end_col = column_index_from_string('H')    # 8



def MergedRowsTitle():
    # Step 1: Collect only merged cells from B to H
    merged_rows = []

    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if min_col == start_col and max_col == end_col:
            merged_rows.append({
                'range': str(merged_range),
                'start_cell': sheet.cell(row=min_row, column=min_col).coordinate,
                'value': sheet.cell(row=min_row, column=min_col).value,
                'row': min_row
            })

    # Step 2: Sort by row number
    merged_rows.sort(key=lambda x: x['row'])
    return merged_rows

def TablesRange():
    # Step 3: Group by consecutive rows
    tables = []
    current_table = []

    for row in merged_rows:
        if not current_table:
            current_table.append(row)
        else:
            prev_row = current_table[-1]['row']
            if row['row'] == prev_row + 1:
                current_table.append(row)
            else:
                tables.append(current_table)
                current_table = [row]
    if current_table:
        tables.append(current_table)
    return tables

# Step 4: Output and collect concatenated values
# def RateTitle():
#     Rate_Title = []
#     for i, table in enumerate(tables):
#         start_row = table[0]['row']
#         if i + 1 < len(tables):
#             next_start_row = tables[i + 1][0]['row']
#             end_row = next_start_row - 1
#         else:
#             end_row = table[-1]['row']  # last group
#
#         # print(f"\nTable {i + 1}: Row {start_row} to {end_row}")
#
#         concatenated_text = ""
#         for row in table:
#             if row['value']:
#                 concatenated_text += str(row['value']).strip() + " "
#         Rate_Title.append(concatenated_text.strip())
#
#         return Rate_Title
def RateTitle():
    Rate_Title_Font11 = []
    Rate_Title_OtherFonts = []

    for i, table in enumerate(tables):
        start_row = table[0]['row']
        if i + 1 < len(tables):
            next_start_row = tables[i + 1][0]['row']
            end_row = next_start_row - 1
        else:
            end_row = table[-1]['row']  # last group

        concatenated_text_font11 = ""
        concatenated_text_other = ""

        for row in table:
            cell = sheet.cell(row=row['row'], column=column_index_from_string('B'))  # cell in col B
            value = cell.value
            font_size = cell.font.sz

            if value:
                if font_size == 7:
                    concatenated_text_font11 += str(value).strip() + " "
                else:
                    concatenated_text_other += str(value).strip() + " "

        Rate_Title_Font11.append(concatenated_text_font11.strip())
        Rate_Title_OtherFonts.append(concatenated_text_other.strip())

    return Rate_Title_Font11, Rate_Title_OtherFonts

# # Step 5: Extract data row just below each table's last merged row
def Rate_ItemsTitles():

    Rate_ItemsTitles = []

    for table in tables:
        last_row = table[-1]['row'] + 1  # row just below the last merged cell
        row_data = []
        for col in range(start_col, end_col + 1):
            cell_value = sheet.cell(row=last_row, column=col).value
            row_data.append(cell_value)
        Rate_ItemsTitles.append(row_data)
        return Rate_ItemsTitles

# # Step 6: Table bounds (A to H) for each group
def TableBounds():

    table_bounds_A_to_H = []
    for i, table in enumerate(tables):

        start_row = table[0]['row']
        if i + 1 < len(tables):
            next_start_row = tables[i + 1][0]['row']
            end_row = next_start_row - 1
        else:
            end_row = table[-1]['row']

        # Define range from A to H
        start_col_letter = get_column_letter(start_col - 1)  # Column A (B is 2)
        end_col_letter = get_column_letter(end_col)
        bounds = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
        table_bounds_A_to_H.append(bounds)
    return table_bounds_A_to_H

# # Step 8: Extract column A values from each table bound (A to H) and segregate
def ColumnAValues():
    colA_values_per_table = []
    colA_integers = []
    colA_non_integers = []

    for i, table in enumerate(tables):
        start_row = table[0]['row']
        if i + 1 < len(tables):
            next_start_row = tables[i + 1][0]['row']
            end_row = next_start_row - 1
        else:
            end_row = table[-1]['row']

        colA_vals = []
        int_vals = []
        non_int_vals = []

        for r in range(start_row, end_row -2):
            val = sheet.cell(row=r, column=1).value  # Column A = 1
            colA_vals.append(val)
            try:
                if isinstance(val, int) or (isinstance(val, str) and val.strip().isdigit()):
                    int_vals.append(int(val))
                else:
                    non_int_vals.append(val)
            except:
                non_int_vals.append(val)

        colA_values_per_table.append(colA_vals)
        colA_integers.append(int_vals)
        colA_non_integers.append(non_int_vals)

    return(colA_values_per_table, colA_integers, colA_non_integers)

# #Determine the inner table bounds for table 1 and 2
def InnertableBounds():
    search_str = vastavikDarrate
    first_inner_bounds = []
    second_inner_bounds = []
    for i, table in enumerate(tables):
        outer_start_row = table[0]['row']
        if i + 1 < len(tables):
            outer_next_start_row = tables[i + 1][0]['row']
            outer_end_row = outer_next_start_row - 1
        else:
            outer_end_row = table[-1]['row']

        # Find the row of the cell containing the search string inside the table bounds (A to H)
        found_row = None
        for r in range(outer_start_row, outer_end_row + 1):
            for c in range(1, end_col + 1):  # columns A(1) to H(end_col=8)
                cell_val = sheet.cell(row=r, column=c).value
                if cell_val and isinstance(cell_val, str) and search_str in cell_val:
                    found_row = r
                    break
            if found_row is not None:
                break

        # Find last merged cell row from B to H in this table
        last_merged_row = table[-1]['row']

        if found_row is None:
            # print(f"Warning: '{search_str}' not found in Table {i + 1}, skipping inner table split.")
            continue

        # First inner table bounds: B to H
        first_inner_start = last_merged_row + 1
        first_inner_end = found_row - 1
        first_inner_bound = f"B{first_inner_start}:H{first_inner_end}"

        # Second inner table bounds: A to H
        second_inner_start = found_row
        second_inner_end = outer_end_row
        second_inner_bound = f"A{second_inner_start}:H{second_inner_end}"

        first_inner_bounds.append(first_inner_bound)
        second_inner_bounds.append(second_inner_bound)
    return first_inner_bounds, second_inner_bounds

# #Step 9: TO get the data of first inner cells, modifying if needed and then
def FirstInnerBounds_Data():
    first_inner_tables_data = []

    for i, table in enumerate(tables):
        outer_start_row = table[0]['row']
        if i + 1 < len(tables):
            outer_next_start_row = tables[i + 1][0]['row']
            outer_end_row = outer_next_start_row - 1
        else:
            outer_end_row = table[-1]['row']

        # Find the row with "jf:tljs b//]6"
        found_row = None
        for r in range(outer_start_row, outer_end_row + 1):
            for c in range(1, end_col + 1):  # Columns A to H
                val = sheet.cell(row=r, column=c).value
                if isinstance(val, str) and vastavikDarrate in val:
                    found_row = r
                    break
            if found_row:
                break

        if not found_row:
            # print(f"Table {i + 1}: 'jf:tljs b//]6' not found. Skipping.")
            first_inner_tables_data.append([])
            continue

        # First inner table bounds
        first_inner_start = table[-1]['row'] + 1
        first_inner_end = found_row - 1

        if first_inner_end < first_inner_start:
            # print(f"Table {i + 1}: Invalid first inner bound. Skipping.")
            continue

        # Extract row-wise data from B to H
        table_rows = []
        for row in range(first_inner_start, first_inner_end + 1):
            row_values = []
            for col in range(start_col, end_col + 1):  # B to H
                cell_val = sheet.cell(row=row, column=col).value
                row_values.append(cell_val)
            table_rows.append(row_values)

        first_inner_tables_data.append(table_rows)

    # ✅ Example output
    # Forward-fill first column if value is None
    for table in first_inner_tables_data:
        last_valid_value = None
        for row in table:
            if row[0] is not None:
                last_valid_value = row[0]
            else:
                row[0] = last_valid_value
    return first_inner_tables_data

# #Strep 10: Extracting the Original Rate, Overhead Value and Total Rate value
def OriginalOverheadTotal_Rate():

    # Define search strings
    search_strings = search_srtings

    # Store the extracted values
    extracted_inner_table_values = []
    rupee_paisa_cells = []
    other_values = []

    matched_cells = set()  # Track (row, col) of matched cells
    rupee_paisa_coords = set()  # Track (row, col) of Rupee-Paisa matches

    for bound_str in second_inner_bounds:
        min_col, min_row, max_col, max_row = range_boundaries(bound_str)
        found_values = []
        Uniquerate = 0.00

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_val = sheet.cell(row=row, column=col).value

                # Match search strings
                if isinstance(cell_val, str) and cell_val.strip() in search_strings:
                    target_col = col + 2
                    target_val = sheet.cell(row=row, column=target_col).value
                    found_values.append(target_val)
                    matched_cells.add((row, target_col))

                # Rupee-Paisa logic
                if min_col < col < max_col:
                    prev_val = sheet.cell(row=row, column=col - 1).value
                    next_val = sheet.cell(row=row, column=col + 1).value
                    if str(prev_val).strip() == rupees and str(next_val).strip() == paisa:
                        Uniquerate = cell_val
                        rupee_paisa_coords.add((row, col))
        rupee_paisa_cells.append(Uniquerate)


        extracted_inner_table_values.append(found_values)

        # Extract other values from the same table that are not already collected
        OtherValues = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if (row, col) in matched_cells or (row, col) in rupee_paisa_coords:
                    continue  # Skip already collected cells
                cell_val = sheet.cell(row=row, column=col).value
                if cell_val in search_strings:
                    continue
                if cell_val not in (None, "",rupees, paisa):
                    OtherValues.append(cell_val)

        other_values.append(OtherValues)
    return {
        "matched_search_values": extracted_inner_table_values,
        "rupee_paisa_values": rupee_paisa_cells,
        "other_values": other_values
    }

# #Strep 11: Extracting the data from the second innter bound and replacing the values from the extracted_inner_table_values and search strings by None
def SecondInnerBounds_Data():

    # Define the search strings again
    search_strings = search_srtings

    # Flatten previously extracted values to use in comparison
    flat_extracted_values = [v for sublist in extracted_inner_table_values for v in sublist if v is not None]

    # Step 1: Extract row-wise data from second inner bounds
    second_inner_tables_data = []

    for bound_str in second_inner_bounds:
        min_col, min_row, max_col, max_row = range_boundaries(bound_str)
        table_data = []

        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value

                # Step 2: Replace search_strings and extracted values with None
                if isinstance(cell_value, str) and cell_value.strip() in search_strings:
                    cell_value = None
                elif cell_value in flat_extracted_values:
                    cell_value = None

                row_data.append(cell_value)
            table_data.append(row_data)

        second_inner_tables_data.append(table_data)
    return second_inner_tables_data

# #Step 12: Determination of Arial of Mangal type of font in column A ie the DUDBC rate reference
def DUDBC_Reference():
    # Lists to store values with specific fonts
    colA_mangal_or_arial = []

    for i, table in enumerate(tables):
        start_row = table[0]['row']
        if i + 1 < len(tables):
            next_start_row = tables[i + 1][0]['row']
            end_row = next_start_row - 1
        else:
            end_row = table[-1]['row']

        font_specific_vals = []

        Reference = '-'
        for r in range(start_row, end_row - 2):
            cell = sheet.cell(row=r, column=1)  # Column A = 1
            val = cell.value
            font_name = cell.font.name
            # font_specific_vals.append(val)

            if font_name in ["Times New Roman"] and val != None :
                Reference = val

        colA_mangal_or_arial.append(Reference)

    return colA_mangal_or_arial

# #Step 13 List to store integer values from Column A per table
def TableNo():
    colA_integers_only = []

    for i, table in enumerate(tables):
        start_row = table[0]['row']
        if i + 1 < len(tables):
            next_start_row = tables[i + 1][0]['row']
            end_row = next_start_row - 1
        else:
            end_row = table[-1]['row']

        for r in range(start_row, end_row - 2):
            cell = sheet.cell(row=r, column=1)  # Column A = 1
            val = cell.value

            try:
                if isinstance(val, int):
                    colA_integers_only.append(val)
                elif isinstance(val, str) and val.strip().isdigit():
                    colA_integers_only.append(int(val))
            except:
                pass
    return colA_integers_only

# #Step 14 : Extract data from the column A and only store the data that doesnot contains ru, numerical data, managal and arial type font data
def Filtered_ColAData():

    # Collect filtered values from Column A (excluding None, integers, Mangal/Arial fonts, and '?=')
    filtered_colA_values_per_table = []

    for i, table in enumerate(tables):
        start_row = table[0]['row']
        if i + 1 < len(tables):
            next_start_row = tables[i + 1][0]['row']
            end_row = next_start_row - 1
        else:
            end_row = table[-1]['row']

        filtered_values = []

        for r in range(start_row, end_row - 2):
            cell = sheet.cell(row=r, column=1)
            val = cell.value
            font_name = cell.font.name

            # Apply all the filters
            if val is None:
                continue
            if isinstance(val, int):
                continue
            if isinstance(val, str) and val.strip().isdigit():
                continue
            if isinstance(val, str) and "?=" in val:
                continue
            if font_name in ["Mangal", "Arial"]:
                continue
            filtered_values.append(val)


        #Test for if the data is previously on the Table NO and Reference or not, before placing it to others
        cleanOthervalues = []
        combined_TablenoReferenceList = colA_integers[i] + [colA_mangal_or_arial[i]] + ['रु.']
        for items in filtered_values:
            # for it_index, item in enumerate(items):
            #     print("Inner", it_index, item)
            if items not in combined_TablenoReferenceList:
                cleanOthervalues.append(items)
        # print(cleanOthervalues)



        filtered_colA_values_per_table.append(cleanOthervalues)
    return filtered_colA_values_per_table



""""Calling Zone"""
merged_rows = MergedRowsTitle()
# print(merged_rows) #Output = [{'range': 'B5:H5', 'start_cell': 'B5', 'value': "!@–#) ;]=dL= uf]nfO{sf] ?v 9fNg] sfo{ xfFufx? sf6L ?vsf] 6'qmf kf/L lgdf{0f :ynaf6 ", 'row': 5}, {'range': 'B6:H6', 'start_cell': 'B6', 'value': '!% dL6/ b"/L;Dd af]sfgL ug]{ sfo{ ;d]t -?vsf] uf]nfO{ hdLg b]lv ! dL6/ dfly gfKg]_ .', 'row': 6}, {'range': 'B7:H7', 'start_cell': 'B7', 'value': 'b/ ljZn]if0fsf] nflu ?v ;+Vof ! lnOPsf]', 'row': 7}

tables = TablesRange()
# print(tables) #Output = [[{'range': 'B5:H5', 'start_cell': 'B5', 'value': "!@–#) ;]=dL= uf]nfO{sf] ?v 9fNg] sfo{ xfFufx? sf6L ?vsf] 6'qmf kf/L lgdf{0f :ynaf6 ", 'row': 5}, {'range': 'B6:H6', 'start_cell': 'B6', 'value': '!% dL6/ b"/L;Dd af]sfgL ug]{ sfo{ ;d]t -?vsf] uf]nfO{ hdLg b]lv ! dL6/ dfly gfKg]_ .', 'row': 6}, {'range': 'B7:H7', 'start_cell': 'B7', 'value': 'b/ ljZn]if0fsf] nflu ?v ;+Vof ! lnOPsf]', 'row': 7}], [{'range': 'B13:H13', 'start_cell': 'B13', 'value': "#!–^) ;]

Rate_Title_Font11, Rate_Title_OtherFonts = RateTitle()
# print(Rate_Title_Font11)  #Output: ['b/ ljZn]if0fsf] nflu ?v ;+Vof ! lnOPsf]', 'b/ ljZn]if0fsf] nflu ?v ;+Vof ! lnOPsf]', 'b/ ljZn]if0fsf] nflu ?v ;+Vof ! lnOPsf]', 'b/ ljZn]if0fsf] nflu ?v ;+Vof ! lnOPsf]', 'b/ ljZn]if0
# print( Rate_Title_OtherFonts)   #Output: ['!@–#) ;]=dL= uf]nfO{sf] ?v 9fNg] sfo{ xfFufx? sf6L ?vsf] 6\'qmf kf/L lgdf{0f :ynaf6 !% dL6/ b"/L;Dd af]sfgL ug]{ sfo{ ;d]t -?vsf] uf]nfO{ hdLg b]lv ! dL6/ dfly gfKg]_ .', '#!–^) ;]=dL

Rate_ItemsTitles = Rate_ItemsTitles()
# print(Rate_ItemsTitles) #Output = [[';|f]t ;fwg', 'tx÷lsl;d', 'kl/df0f', 'PsfO{', 'b/ k|lt PsfO{', '/sd', 'k|To]s ;|f]t ;fwgsf] hDdf']]

table_bounds_A_to_H = TableBounds()
# print(table_bounds_A_to_H)  #Output = ['A5:H12', 'A13:H21', 'A22:H29', 'A30:H37', 'A38:H45', 'A46:H53', 'A54:H60', 'A61:H67'

colA_values_per_table, colA_integers, colA_non_integers = ColumnAValues()
# print(colA_values_per_table)  #Output = [[1, 'A 1', None, None, None]]
# print(colA_integers)  #Output = [[1]]
# print(colA_non_integers)  #Output = [['A 1', None, None, None]]

first_inner_bounds, second_inner_bounds = InnertableBounds()
# print(first_inner_bounds, second_inner_bounds) #Output = ['B8:H9', 'B16:H17', 'B25:H26', 'B33:H34', 'B41:H42', 'B49:H50', 'B56:H57', 'B63:H64', 'B70:H73', 'B80:H87', 'B94:H103', 'B113:H115', 'B122:H124', 'B131:H133', 'B140:H143',

first_inner_tables_data = FirstInnerBounds_Data()
# print(first_inner_tables_data) #Output : [[[';|f]t ;fwg', 'tx÷lsl;d', 'kl/df0f', 'PsfO{', 'b/ k|lt PsfO{', '/sd', 'k|To]s ;|f]t ;fwgsf] hDdf'], ['>lds', 'Hofld', 0.13, 'hjfg', 863, 112.19, 112.19]], [[';|f]t ;fwg', 't

InnerTableResults = OriginalOverheadTotal_Rate()
# Access the different sets of extracted data
extracted_inner_table_values = InnerTableResults["matched_search_values"]
rupee_paisa_values = InnerTableResults["rupee_paisa_values"]
InnerTable_other_values = InnerTableResults["other_values"]
# print(len(tables), len(first_inner_bounds), len(second_inner_bounds))

# print(extracted_inner_table_values) #Output : [[112.19, 16.82, 129.01], [336.57, 50.480000000000004, 387.05], [845.74, 126.86, 972.6], [345.2, 51.78, 396.98], [457.39, 68.60000000000001, 525.99], [2174.76, 326.21, 2500.9700000000003


# second_inner_tables_data = SecondInnerBounds_Data()
# # print(second_inner_tables_data ) #Output : [[[None, None, None, None, None, None, None, None], [None, 'b/ k|lt ! ?vsf]', None, None, None, None, None, None], ['?=', None, 'k}=', None, None, None, None, None]], [[None, None, None, None, None, None,

colA_mangal_or_arial = DUDBC_Reference()
# print(colA_mangal_or_arial) #Output : ['A 1', 'A 1', 'A 1', 'A 2', 'A 2', 'A 2', 'A3', 'A4', 'A5', 'A6', 'A7', 'B 1', 'B 1', 'B 1', 'B 2', 'B 2', 'B 3', 'B 3', 'B 4', 'B5', 'B6', 'B7', 'B7', 'C 1', 'C 1', 'C 1', 'C 1', 'C 1', 'C 1', 'C 2', 'C 2', 'C 2', 'C 2', 'C 2', 'C 2', 'C 2', 'C3', 'C4', 'C6', 'C6', 'C6', 'C6', 'C6', 'C9', 'C10', 'C13', 'C13', 'D1', 'D1', 'D1', 'D2', 'D2', 'D2', 'D3', 'D3', 'D4', 'D4', 'D5', 'D6', 'D7', 'D9', '-', 'E1', 'E2', 'E3', 'E4', 'E5', 'E6', 'E12', 'E13', 'F1', 'F1', 'F1', 'F4', 'F4', 'F6', 'F6', 'F6', 'F6', 'F4', 'F7', 'F8', 'F9', 'F10', 'F11', 'F14', 'F14', 'F15', 'F19', 'F20', 'F23', 'F21', 'G2', 'G3', 'G3', 'G4', 'G4', 'G4', 'G4', 'G4', 'G4', 'G5', 'G5', 'G5', '-', '-'

# colA_integers_only = TableNo()
filtered_colA_values_per_table = Filtered_ColAData()
# print(filtered_colA_values_per_table) #Output: [[], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], ['% -@_ s'], ['% -@_ v'], ['% -@_ s'], ['% -@_ v'], ['% -@_ s'],



"""Writing Zone"""
def IntegratedInfo():
    IntegratedData = {}

    for data_index, table_group in enumerate(tables):
        #Track the point where the dictionary writing needs to be end
        table_Nos = colA_integers[data_index]
        TableNo = 0
        try:
            if isinstance(table_Nos[0], int):
                TableNo = table_Nos[0]
        except:
            pass

        if TableNo == 326:
            break

        IntegratedData[data_index+1] = {"Title_Section":{"Title": [], "Note": []},
                                 "References": {"Table No": [], "Reference" : [], "Others": []},
                                 "First Inner table": {"Title": [], "Manpower": [], "Materials": [], "Machines": [], "Others": []},
                                 "Second Inner table": {"Original Rate": [], "Overhead": [], "Total Rate": [], "Unit Rate":[], "Others": []}}


        #Entry of titles and notes_All under the title section
        Title = Rate_Title_OtherFonts[data_index]
        Title_note =  Rate_Title_Font11[data_index]
        IntegratedData[data_index + 1]["Title_Section"]["Title"].append(Title)
        IntegratedData[data_index + 1]["Title_Section"]["Note"].append(Title_note)

        #All entries under the References
        IntegratedData[data_index + 1]["References"]["Table No"].append(TableNo)

        DudbcReference = colA_mangal_or_arial[data_index]
        IntegratedData[data_index + 1]["References"]["Reference"].append(DudbcReference)

        Ref_Others = filtered_colA_values_per_table[data_index]
        # print(Ref_Others, IntegratedData[data_index + 1]["References"]["Others"])
        IntegratedData[data_index + 1]["References"]["Others"] = Ref_Others
            # .append(Ref_Others)

        #All entries under the first Inner table data
        FirstTable_Title = first_inner_tables_data[data_index][0]
        IntegratedData[data_index + 1]["First Inner table"]["Title"].append(FirstTable_Title)

        for tabledata in first_inner_tables_data[data_index][1:]:
            # print(tabledata[0])
            # codes = [ "श्रमिक","श्रमिक ", "निर्माण सामग्री","स्रोत साधन","निर्माण सामग्री लगाउने ज्याला समेत",   "यान्त्रिक उपकरण" ]
            # if tabledata[0] not in codes:
            #     print(f'---{tabledata[0]}---', DudbcReference)


            if tabledata[0] in shramik:
                IntegratedData[data_index + 1]["First Inner table"]["Manpower"].append(tabledata)
            elif tabledata[0] in nirmaSamagri:
                IntegratedData[data_index + 1]["First Inner table"]["Materials"].append(tabledata)
            elif tabledata[0] in yantrikUpakaran:
                IntegratedData[data_index + 1]["First Inner table"]["Machines"].append(tabledata)
            else:
                IntegratedData[data_index + 1]["First Inner table"]["Others"].append(tabledata)

            # if tabledata[0] == "श्रमिक":
            #     IntegratedData[data_index + 1]["First Inner table"]["Manpower"].append(tabledata)
            # elif tabledata[0] == "श्रमिक ":
            #     IntegratedData[data_index + 1]["First Inner table"]["Manpower"].append(tabledata)
            # elif tabledata[0] == "निर्माण सामग्री":
            #     IntegratedData[data_index + 1]["First Inner table"]["Materials"].append(tabledata)
            # elif tabledata[0] == "स्रोत साधन":
            #     IntegratedData[data_index + 1]["First Inner table"]["Materials"].append(tabledata)
            # elif tabledata[0] == "निर्माण सामग्री लगाउने ज्याला समेत":
            #     IntegratedData[data_index + 1]["First Inner table"]["Materials"].append(tabledata)
            # elif tabledata[0] == "यान्त्रिक उपकरण":
            #     IntegratedData[data_index + 1]["First Inner table"]["Machines"].append(tabledata)
            # elif tabledata[0] == "यान्त्रिक उपकरण":
            #     IntegratedData[data_index + 1]["First Inner table"]["Machines"].append(tabledata)
            # else:
            #     IntegratedData[data_index + 1]["First Inner table"]["Others"].append(tabledata)


            # if tabledata[0] == ">lds":
            #     IntegratedData[data_index + 1]["First Inner table"]["Manpower"].append(tabledata)
            # elif tabledata[0] == "lgdf{0f ;fdu|L":
            #     IntegratedData[data_index + 1]["First Inner table"]["Materials"].append(tabledata)
            # elif tabledata[0] == ">oflGqs pks/0f":
            #     IntegratedData[data_index + 1]["First Inner table"]["Machines"].append(tabledata)
            # elif tabledata[0] == "oflGqs pks/0f":
            #     IntegratedData[data_index + 1]["First Inner table"]["Machines"].append(tabledata)
            # else:
            #     IntegratedData[data_index + 1]["First Inner table"]["Others"].append(tabledata)

        #All entries under the first Inner table data
        # print(table_Nos, extracted_inner_table_values[data_index], rupee_paisa_values[data_index], InnerTable_other_values[data_index], second_inner_bounds[data_index])
        OriginalCost = extracted_inner_table_values[data_index][0]
        OverheadValue = extracted_inner_table_values[data_index][1]
        TotalCost = extracted_inner_table_values[data_index][2]
        UnitRate = rupee_paisa_values[data_index]
        OtherNotes = InnerTable_other_values[data_index]
        IntegratedData[data_index + 1]["Second Inner table"]["Original Rate"].append(OriginalCost)
        IntegratedData[data_index + 1]["Second Inner table"]["Overhead"].append(OverheadValue)
        IntegratedData[data_index + 1]["Second Inner table"]["Total Rate"].append(TotalCost)
        IntegratedData[data_index + 1]["Second Inner table"]["Unit Rate"].append(UnitRate)
        IntegratedData[data_index + 1]["Second Inner table"]["Others"] = OtherNotes
            # .append(OtherNotes)


    return IntegratedData





IntegratedData = IntegratedInfo()
print(IntegratedData)
# for key, value in IntegratedData.items():
#     print(key)
#     print(value)
#     print(value['References'])






def transform_and_write_all_tables(IntegratedData, output_filename):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set column widths
    column_widths = {'A': 15, 'B': 15, 'C': 10, 'D': 10, 'E': 10,
                     'F': 15, 'G': 15, 'H': 10, 'I': 10, 'J': 15,
                     'K': 15, 'L': 10, 'M': 10, 'N': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    current_row = 1  # Track current row position

    for table_idx, table_data in enumerate(IntegratedData):
        if table_data == 322:
            break
        # print(table_data)
        table_data = IntegratedData[table_data]
        # print( "Helo NEpaa", table_data['Title_Section']['Title'])
        if table_idx > 0:  # Add spacing between tables (except before first table)
            current_row += 3

        # Write description section
        ws.merge_cells(f'A{current_row}:N{current_row}')
        description = table_data['Title_Section']['Title'][0]
        ws[f'A{current_row}'] = f"Description of works: \t\t{description}"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'A{current_row}'].font = Font(bold=True)

        # Add the unit rate at the end of description
        # others_second_table = table_data['Second Inner table']['Others'][0][0]
        # ws[f'N{current_row}'] = others_second_table
        # ws[f'N{current_row}'].alignment = Alignment(horizontal='right')
        current_row += 1

        # Add Spec. cl. No
        ws.merge_cells(f'A{current_row}:B{current_row}')
        reference = table_data['References']['Reference'][0]
        ws[f'A{current_row}'] = f"Spec. cl. No: \t{reference}"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        # Add Norms No header
        ws.merge_cells(f'A{current_row}:A{current_row + 1}')
        ws[f'A{current_row}'] = "Norms No"
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add headers for each section
        ws.merge_cells(f'B{current_row}:F{current_row}')
        ws[f'B{current_row}'] = "श्रमिक (क)"
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'G{current_row}:K{current_row}')
        ws[f'G{current_row}'] = "निर्माण सामग्री (ख)"
        ws[f'G{current_row}'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'L{current_row}:N{current_row}')
        ws[f'L{current_row}'] = "यान्त्रिक उपकरण (ग)"
        ws[f'L{current_row}'].alignment = Alignment(horizontal='center')

        # Add column headers
        column_headers = ['Type', 'Unit', 'Qty.', 'Rate', 'Amount'] * 2 + ['Type', 'Unit', 'Qty.', 'Rate', 'Amount']
        for i, header in enumerate(column_headers[:5], start=2):
            ws.cell(row=current_row + 1, column=i, value=header)
        for i, header in enumerate(column_headers[5:10], start=7):
            ws.cell(row=current_row + 1, column=i, value=header)
        for i, header in enumerate(column_headers[10:], start=12):
            ws.cell(row=current_row + 1, column=i, value=header)

        current_row += 2

        # Extract data from First Inner table
        others_data = table_data['First Inner table']['Others']

        # Separate data by category
        manpower_data = []
        materials_data = []
        machines_data = []

        for row in others_data:
            if row[0] == 'श्रमिक':
                manpower_data.append(row[1:])
            elif row[0] == 'निर्माण सामग्री':
                materials_data.append(row[1:])
            elif row[0] == 'यान्त्रिक उपकरण':
                machines_data.append(row[1:])

        # Determine how many rows we'll need for this table
        data_rows = max(len(manpower_data), len(materials_data), len(machines_data))
        start_data_row = current_row

        # Write manpower data
        norms_no = table_data['References']['Reference'][0]
        for i, row in enumerate(manpower_data):
            ws.cell(row=start_data_row + i, column=1, value=norms_no if i == 0 else "")
            for j, value in enumerate(row[:5]):
                ws.cell(row=start_data_row + i, column=j + 2, value=value)

        # Write materials data
        for i, row in enumerate(materials_data):
            for j, value in enumerate(row[:5]):
                ws.cell(row=start_data_row + i, column=j + 7, value=value)

        # Write machines data
        for i, row in enumerate(machines_data):
            for j, value in enumerate(row[:5]):
                ws.cell(row=start_data_row + i, column=j + 12, value=value)

        current_row = start_data_row + data_rows

        # Add total rows
        total_ka_row = current_row
        ws.merge_cells(f'D{total_ka_row}:F{total_ka_row}')
        ws[f'D{total_ka_row}'] = "जम्मा क ="
        if manpower_data and len(manpower_data[-1]) > 5:
            ws[f'G{total_ka_row}'] = manpower_data[-1][-1]  # Last manpower total
            ws[f'G{total_ka_row}'].number_format = '#,##0.00'

        ws.merge_cells(f'I{total_ka_row}:K{total_ka_row}')
        ws[f'I{total_ka_row}'] = "जम्मा ख ="
        if materials_data and len(materials_data[-1]) > 5:
            ws[f'L{total_ka_row}'] = materials_data[-1][-1]  # Last materials total
            ws[f'L{total_ka_row}'].number_format = '#,##0.00'

        ws.merge_cells(f'M{total_ka_row}:N{total_ka_row}')
        ws[f'M{total_ka_row}'] = "जम्मा ग ="
        if machines_data and len(machines_data[0]) > 5:
            # ws[f'N{total_ka_row}'] = machines_data[0][-1]  # Machines total
            ws[f'N{total_ka_row}'].number_format = '#,##0.00'

        # Add actual rate row
        actual_rate_row = total_ka_row + 1
        original_rate = table_data['Second Inner table']['Original Rate'][0]
        overhead = table_data['Second Inner table']['Overhead'][0]
        total_rate = table_data['Second Inner table']['Total Rate'][0]

        ws.merge_cells(f'D{actual_rate_row}:F{actual_rate_row}')
        ws[f'D{actual_rate_row}'] = f"वास्तविक दररेट( क + ख + ग) = "
        ws[f'G{actual_rate_row}'] = original_rate
        ws[f'G{actual_rate_row}'].number_format = '#,##0.00'

        ws.merge_cells(f'I{actual_rate_row}:K{actual_rate_row}')
        ws[f'I{actual_rate_row}'] = "१५% ठेकेदार ओभरहेड"
        # ws[f'L{actual_rate_row}'] = total_rate - original_rate
        ws[f'L{actual_rate_row}'].number_format = '#,##0.00'

        ws.merge_cells(f'M{actual_rate_row}:N{actual_rate_row}')
        # ws[f'M{actual_rate_row}'] = "जम्मा दर रेट"
        # ws[f'N{actual_rate_row}'] = total_rate
        # ws[f'N{actual_rate_row}'].number_format = '#,##0.00'

        # Add unit rate row
        unit_rate_row = actual_rate_row + 1
        ws.merge_cells(f'M{unit_rate_row}:N{unit_rate_row}')
        ws[f'M{unit_rate_row}'] = "दर प्रति रुखको"
        # ws[f'N{unit_rate_row}'] = table_data['Second Inner table']['Unit Rate'][0]
        ws[f'N{unit_rate_row}'].number_format = '#,##0.00'

        current_row = unit_rate_row + 1

    # Save the workbook
    wb.save(output_filename)


# Example usage:
# transform_and_write_all_tables(IntegratedData, "all_tables_output.xlsx")




def FontMapping(FilePath, Worksheet):
    # Initialize the font mapper with mapping rules
    # mapper = npttf2utf.FontMapper("npttf2utf/map.json")  # Adjust path if necessary Download it from: https://github.com/casualsnek/npttf2utf/blob/main/src/npttf2utf/map.json
    mapper = npttf2utf.FontMapper(r"C:\Users\Acer\Downloads\npttf2utf-main\npttf2utf-main\src\npttf2utf\map.json")

    #Conversion of Preeti font to Kalimati font under the selected rows of provided Worksheet
    input_file =  FilePath # Your input Excel file
    output_file = "rate_analysis_unicode.xlsx"  # Output file
    wb = openpyxl.load_workbook(input_file)
    ws = wb[Worksheet]

    # Process cells with font set to "Preeti" only
    for row in ws.iter_rows(min_row=1, max_row=3720, min_col=1, max_col=8):
        for cell in row:
            if isinstance(cell.value, str):
                font_name = cell.font.name
                if font_name and "Preeti" in font_name:
                    try:
                        converted = mapper.map_to_unicode(cell.value, from_font="Preeti")
                        cell.value = converted
                        # Optionally change font to Kalimati to reflect Unicode encoding
                        original_font_size = cell.font.size
                        set_font_size = 9
                        if original_font_size == 11:
                            set_font_size = 7
                        cell.font = Font(name="Kalimati", size=set_font_size)
                    except Exception as e:
                        print(f"Error at cell {cell.coordinate}: {e}")

    # Save the updated file
    wb.save(output_file)
    print(f"Conversion complete. Saved as '{output_file}'")

# Load the Excel workbook and target worksheet
FilePath = r"C:\Users\Acer\Desktop\Software_Input.xlsx"
Worksheet = "Analysis"
# FontMapping(FilePath, Worksheet)


