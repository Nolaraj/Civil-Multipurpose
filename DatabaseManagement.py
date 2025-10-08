import sqlite3
from sqlite3 import Error
from typing import Union, Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment
import openpyxl as op
from itertools import zip_longest
from openpyxl.styles import Border, Side
import xlwings as xw
from PyPDF2 import PdfMerger
import sys


def resourece_path(rel_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, rel_path)
def DUDBC_RateWriter(databasepath):

    def create_connection(db_file):
        """ Create a database connection to a SQLite database """
        conn = None
        try:
            conn = sqlite3.connect(db_file)
            return conn
        except Error as e:
            print(e)
        return conn


    def create_tables(conn):
        """ Create all tables needed for the hierarchical data structure """

        # Drop tables if they exist (for development)
        drop_tables = [
            "Second_Inner_Data", "Second_Inner_Table",
            "Others", "Machines", "Materials", "Manpower",
            "First_Inner_Title", "First_Inner_Table",
            "Reference_Data_Items", "Reference_Data",
            "Title_Section_Data", "Title_Section",
            "Root"
        ]

        try:
            c = conn.cursor()
            for table in drop_tables:
                c.execute(f"DROP TABLE IF EXISTS {table}")
        except Error as e:
            print(f"Error dropping tables: {e}")

        # Create fresh tables
        sql_create_root_table = """CREATE TABLE IF NOT EXISTS Root (
                                    id INTEGER PRIMARY KEY
                                );"""

        sql_create_title_section_table = """CREATE TABLE IF NOT EXISTS Title_Section (
                                            id INTEGER PRIMARY KEY,
                                            root_id INTEGER NOT NULL,
                                            FOREIGN KEY (root_id) REFERENCES Root (id)
                                        );"""

        sql_create_title_section_data = """CREATE TABLE IF NOT EXISTS Title_Section_Data (
                                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                                            title_section_id INTEGER NOT NULL,
                                            key TEXT NOT NULL,
                                            value TEXT NOT NULL,
                                            FOREIGN KEY (title_section_id) REFERENCES Title_Section (id)
                                        );"""

        sql_create_references_table = """CREATE TABLE IF NOT EXISTS Reference_Data (
                                        id INTEGER PRIMARY KEY,
                                        root_id INTEGER NOT NULL,
                                        FOREIGN KEY (root_id) REFERENCES Root (id)
                                    );"""

        sql_create_references_data = """CREATE TABLE IF NOT EXISTS Reference_Data_Items (
                                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                                        reference_data_id INTEGER NOT NULL,
                                        key TEXT NOT NULL,
                                        value TEXT NOT NULL,
                                        FOREIGN KEY (reference_data_id) REFERENCES Reference_Data (id)
                                    );"""

        sql_create_first_inner_table = """CREATE TABLE IF NOT EXISTS First_Inner_Table (
                                        id INTEGER PRIMARY KEY,
                                        root_id INTEGER NOT NULL,
                                        FOREIGN KEY (root_id) REFERENCES Root (id)
                                    );"""

        sql_create_first_inner_title = """CREATE TABLE IF NOT EXISTS First_Inner_Title (
                                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                                        first_inner_id INTEGER NOT NULL,
                                        value TEXT NOT NULL,
                                        FOREIGN KEY (first_inner_id) REFERENCES First_Inner_Table (id)
                                    );"""

        sql_create_manpower_table = """CREATE TABLE IF NOT EXISTS Manpower (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    first_inner_id INTEGER NOT NULL,
                                    category TEXT,
                                    level_type TEXT,
                                    quantity REAL,
                                    unit TEXT,
                                    rate_per_unit REAL,
                                    amount REAL,
                                    total_per_resource REAL,
                                    FOREIGN KEY (first_inner_id) REFERENCES First_Inner_Table (id)
                                );"""

        sql_create_materials_table = """CREATE TABLE IF NOT EXISTS Materials (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    first_inner_id INTEGER NOT NULL,
                                    category TEXT,
                                    level_type TEXT,
                                    quantity REAL,
                                    unit TEXT,
                                    rate_per_unit REAL,
                                    amount REAL,
                                    total_per_resource REAL,
                                    FOREIGN KEY (first_inner_id) REFERENCES First_Inner_Table (id)
                                );"""

        sql_create_machines_table = """CREATE TABLE IF NOT EXISTS Machines (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    first_inner_id INTEGER NOT NULL,
                                    category TEXT,
                                    level_type TEXT,
                                    quantity REAL,
                                    unit TEXT,
                                    rate_per_unit REAL,
                                    amount REAL,
                                    total_per_resource REAL,
                                    FOREIGN KEY (first_inner_id) REFERENCES First_Inner_Table (id)
                                );"""

        sql_create_others_table = """CREATE TABLE IF NOT EXISTS Others (
                                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                                    first_inner_id INTEGER NOT NULL,
                                    category TEXT,
                                    level_type TEXT,
                                    quantity REAL,
                                    unit TEXT,
                                    rate_per_unit REAL,
                                    amount REAL,
                                    total_per_resource REAL,
                                    FOREIGN KEY (first_inner_id) REFERENCES First_Inner_Table (id)
                                );"""

        sql_create_second_inner_table = """CREATE TABLE IF NOT EXISTS Second_Inner_Table (
                                        id INTEGER PRIMARY KEY,
                                        root_id INTEGER NOT NULL,
                                        FOREIGN KEY (root_id) REFERENCES Root (id)
                                    );"""

        sql_create_second_inner_data = """CREATE TABLE IF NOT EXISTS Second_Inner_Data (
                                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                                        second_inner_id INTEGER NOT NULL,
                                        key TEXT NOT NULL,
                                        value TEXT NOT NULL,
                                        FOREIGN KEY (second_inner_id) REFERENCES Second_Inner_Table (id)
                                    );"""

        try:
            c = conn.cursor()
            c.execute(sql_create_root_table)
            c.execute(sql_create_title_section_table)
            c.execute(sql_create_title_section_data)
            c.execute(sql_create_references_table)
            c.execute(sql_create_references_data)
            c.execute(sql_create_first_inner_table)
            c.execute(sql_create_first_inner_title)
            c.execute(sql_create_manpower_table)
            c.execute(sql_create_materials_table)
            c.execute(sql_create_machines_table)
            c.execute(sql_create_others_table)
            c.execute(sql_create_second_inner_table)
            c.execute(sql_create_second_inner_data)
        except Error as e:
            print(e)


    def insert_or_update_data(conn, data):
        """ Insert or update the hierarchical data in the database """

        for root_id, root_data in data.items():
            # Check if root exists
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM Root WHERE id = ?", (root_id,))
            exists = cursor.fetchone()

            if exists:
                # Delete existing data for this root to avoid duplicates
                delete_root_data(conn, root_id)

            # Insert root
            conn.execute("INSERT INTO Root (id) VALUES (?)", (root_id,))

            # Insert Title_Section
            title_section = root_data.get('Title_Section', {})
            conn.execute("INSERT INTO Title_Section (id, root_id) VALUES (?, ?)",
                         (root_id, root_id))

            for key, values in title_section.items():
                for value in values:
                    if value:  # Only insert non-empty values
                        conn.execute("INSERT INTO Title_Section_Data (title_section_id, key, value) VALUES (?, ?, ?)",
                                     (root_id, key, value))

            # Insert Reference_Data
            references = root_data.get('References', {})
            conn.execute("INSERT INTO Reference_Data (id, root_id) VALUES (?, ?)",
                         (root_id, root_id))

            for key, values in references.items():
                for value in values:
                    if value or value == 0:  # Allow zero values
                        conn.execute("INSERT INTO Reference_Data_Items (reference_data_id, key, value) VALUES (?, ?, ?)",
                                     (root_id, key, str(value)))

            # Insert First Inner Table
            first_inner = root_data.get('First Inner table', {})
            conn.execute("INSERT INTO First_Inner_Table (id, root_id) VALUES (?, ?)",
                         (root_id, root_id))

            # Insert Title rows - only non-empty values
            for title_row in first_inner.get('Title', []):
                for value in title_row:
                    if value:  # Only insert non-empty values
                        conn.execute("INSERT INTO First_Inner_Title (first_inner_id, value) VALUES (?, ?)",
                                     (root_id, value))

            # Insert Manpower data
            for row in first_inner.get('Manpower', []):
                if row:  # Only insert if row exists
                    conn.execute("""INSERT INTO Manpower 
                                (first_inner_id, category, level_type, quantity, unit, rate_per_unit, amount, total_per_resource) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                                 (root_id, row[0], row[1], row[2], row[3], row[4], row[5], row[6]))

            # Insert Materials data
            for row in first_inner.get('Materials', []):
                if row:  # Only insert if row exists
                    conn.execute("""INSERT INTO Materials 
                                (first_inner_id, category, level_type, quantity, unit, rate_per_unit, amount, total_per_resource) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                                 (root_id, row[0], row[1], row[2], row[3], row[4], row[5], row[6]))

            # Insert Machines data
            for row in first_inner.get('Machines', []):
                if row:  # Only insert if row exists
                    conn.execute("""INSERT INTO Machines 
                                (first_inner_id, category, level_type, quantity, unit, rate_per_unit, amount, total_per_resource) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                                 (root_id, row[0], row[1], row[2], row[3], row[4], row[5], row[6]))

            # Insert Others data
            for row in first_inner.get('Others', []):
                if row:  # Only insert if row exists
                    conn.execute("""INSERT INTO Others 
                                (first_inner_id, category, level_type, quantity, unit, rate_per_unit, amount, total_per_resource) 
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                                 (root_id, row[0], row[1], row[2], row[3], row[4], row[5], row[6]))

            # Insert Second Inner Table
            second_inner = root_data.get('Second Inner table', {})
            conn.execute("INSERT INTO Second_Inner_Table (id, root_id) VALUES (?, ?)",
                         (root_id, root_id))

            for key, values in second_inner.items():
                for value in values:
                    if value or value == 0:  # Allow zero values
                        conn.execute("INSERT INTO Second_Inner_Data (second_inner_id, key, value) VALUES (?, ?, ?)",
                                     (root_id, key, str(value)))
        conn.commit()


    def delete_root_data(conn, root_id):
        """ Delete all data associated with a root_id """
        try:
            # Delete in reverse order of foreign key dependencies
            conn.execute(
                "DELETE FROM Second_Inner_Data WHERE second_inner_id IN (SELECT id FROM Second_Inner_Table WHERE root_id = ?)",
                (root_id,))
            conn.execute("DELETE FROM Second_Inner_Table WHERE root_id = ?", (root_id,))

            conn.execute(
                "DELETE FROM Manpower WHERE first_inner_id IN (SELECT id FROM First_Inner_Table WHERE root_id = ?)",
                (root_id,))
            conn.execute(
                "DELETE FROM Materials WHERE first_inner_id IN (SELECT id FROM First_Inner_Table WHERE root_id = ?)",
                (root_id,))
            conn.execute(
                "DELETE FROM Machines WHERE first_inner_id IN (SELECT id FROM First_Inner_Table WHERE root_id = ?)",
                (root_id,))
            conn.execute("DELETE FROM Others WHERE first_inner_id IN (SELECT id FROM First_Inner_Table WHERE root_id = ?)",
                         (root_id,))
            conn.execute(
                "DELETE FROM First_Inner_Title WHERE first_inner_id IN (SELECT id FROM First_Inner_Table WHERE root_id = ?)",
                (root_id,))
            conn.execute("DELETE FROM First_Inner_Table WHERE root_id = ?", (root_id,))

            conn.execute(
                "DELETE FROM Reference_Data_Items WHERE reference_data_id IN (SELECT id FROM Reference_Data WHERE root_id = ?)",
                (root_id,))
            conn.execute("DELETE FROM Reference_Data WHERE root_id = ?", (root_id,))

            conn.execute(
                "DELETE FROM Title_Section_Data WHERE title_section_id IN (SELECT id FROM Title_Section WHERE root_id = ?)",
                (root_id,))
            conn.execute("DELETE FROM Title_Section WHERE root_id = ?", (root_id,))

            conn.execute("DELETE FROM Root WHERE id = ?", (root_id,))
        except Error as e:
            print(f"Error deleting root data: {e}")


    def main():
        database = databasepath

        # Sample data (you provided)
        import InputData
        MappedData = InputData.mapped_Data

        # Create a database connection
        conn = create_connection(database)

        if conn is not None:
            # Create tables (will drop existing ones)
            create_tables(conn)

            # Insert data
            insert_or_update_data(conn, MappedData)

            # Close connection
            conn.close()
        else:
            print("Error! cannot create the database connection.")


    if __name__ == '__main__':
        main()





@dataclass
class SearchResult:
    root_id: int
    matched_field: str
    matched_value: str
    extracted_data: Dict[str, Any]
class SuperDataExtractor:
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn = None

    def __enter__(self):
        self.connect()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def connect(self):
        try:
            self.conn = sqlite3.connect(self.db_path)
            self.conn.row_factory = sqlite3.Row
        except sqlite3.Error as e:
            raise ConnectionError(f"Database connection error: {e}")

    def close(self):
        if self.conn:
            self.conn.close()

    def smart_search(
            self,
            title_pattern: Optional[str] = None,
            content_pattern: Optional[str] = None,
            content_column: Optional[str] = None,
            extract_fields: Optional[List[str]] = None
    ) -> List[SearchResult]:
        """
        Flexible search across titles or table content
        Args:
            title_pattern: String to search in titles (None to skip title search)
            content_pattern: String to search in content (None to skip content search)
            content_column: Specific column to search in (None for any column)
            extract_fields: List of specific fields to extract ('Table_No', 'Overhead', etc.)
        Returns:
            List of SearchResult objects with matched data
        """
        results = []

        # Search by title if pattern provided
        if title_pattern:
            title_matches = self._search_titles(title_pattern)
            for match in title_matches:
                extracted = self._extract_specific_fields(match['root_id'], extract_fields)
                results.append(SearchResult(
                    root_id=match['root_id'],
                    matched_field='Title',
                    matched_value=match['title'],
                    extracted_data=extracted
                ))

        # Search by content if pattern provided
        if content_pattern:
            content_matches = self._search_content(content_pattern, content_column)
            for match in content_matches:
                # Avoid duplicates if same root_id matched by both title and content
                if not any(r.root_id == match['root_id'] for r in results):
                    extracted = self._extract_specific_fields(match['root_id'], extract_fields)
                    results.append(SearchResult(
                        root_id=match['root_id'],
                        matched_field=match['column'],
                        matched_value=match['matched_value'],
                        extracted_data=extracted
                    ))

        return results

    def _search_titles(self, pattern: str) -> List[Dict]:
        """Search for pattern in table titles"""
        query = """
            SELECT tsd.title_section_id as root_id, tsd.value as title
            FROM Title_Section_Data tsd
            WHERE tsd.key = 'Title' AND tsd.value LIKE ?
        """
        return self._execute_query(query, (f'%{pattern}%',))

    def _search_content(self, pattern: str, column: Optional[str] = None) -> List[Dict]:
        """
        Search for pattern in table content
        Searches: Manpower, Materials, Machines, Others tables
        """
        tables = ['Manpower', 'Materials', 'Machines', 'Others']
        results = []

        for table in tables:
            # Determine which columns to search
            if column:
                columns = [column]
            else:
                # Get all text columns for the table
                columns = self._get_text_columns(table)

            for col in columns:
                query = f"""
                    SELECT ft.root_id, ? as column, {col} as matched_value
                    FROM {table} t
                    JOIN First_Inner_Table ft ON t.first_inner_id = ft.id
                    WHERE {col} LIKE ?
                """
                matches = self._execute_query(query, (col, f'%{pattern}%'))
                results.extend(matches)

        return results

    def _get_text_columns(self, table: str) -> List[str]:
        """Get all text-type columns for a table"""
        # These are the text columns in our schema
        table_columns = {
            'Manpower': ['category', 'level_type', 'unit'],
            'Materials': ['category', 'level_type', 'unit'],
            'Machines': ['category', 'level_type', 'unit'],
            'Others': ['category', 'level_type', 'unit']
        }
        return table_columns.get(table, [])

    def _extract_specific_fields(self, root_id: int, fields: Optional[List[str]]) -> Dict[str, Any]:
        """Extract only requested fields from a table"""
        if not fields:
            return {}

        result = {}

        for field in fields:
            if field == 'Table_No':
                ref_data = self._execute_query("""
                    SELECT value FROM Reference_Data_Items
                    WHERE reference_data_id = ? AND key = 'Table No'
                """, (root_id,))
                if ref_data:
                    result['Table_No'] = ref_data[0]['value']

            elif field == 'Overhead':
                overhead_data = self._execute_query("""
                    SELECT value FROM Second_Inner_Data
                    WHERE second_inner_id = ? AND key = 'Overhead'
                """, (root_id,))
                if overhead_data:
                    result['Overhead'] = self._convert_value(overhead_data[0]['value'])

            elif field == 'Title':
                title_data = self._execute_query("""
                    SELECT value FROM Title_Section_Data
                    WHERE title_section_id = ? AND key = 'Title'
                """, (root_id,))
                if title_data:
                    result['Title'] = title_data[0]['value']

            # Add more field extractors as needed

        return result

    def _execute_query(self, query: str, params: Tuple = ()) -> List[Dict]:
        """Execute query with parameters and return results as dicts"""
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, params)
            return [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error as e:
            raise ValueError(f"Query execution error: {e}")

    def _convert_value(self, value: str) -> Union[int, float, str]:
        """Convert string values to appropriate Python types"""
        if value is None:
            return None
        try:
            return float(value) if '.' in value else int(value)
        except ValueError:
            return value

    def search_resources(
            self,
            resource_type: str,  # 'all', 'labour', 'materials', 'machines'
            search_column: str,  # Column name to search in
            search_value: str,  # Value to search for
            exact_match: bool = False
    ) -> List[Dict[str, Any]]:
        """
        Search across specified resource types and return matching tables
        Args:
            resource_type: Type of resource to search
            search_column: Column name to search in
            search_value: Value to search for
            exact_match: Whether to require exact match
        Returns:
            List of matching tables with basic info
        """
        # Map resource types to tables
        resource_tables = {
            'labour': ['Manpower'],
            'materials': ['Materials'],
            'machines': ['Machines'],
            'all': ['Manpower', 'Materials', 'Machines', 'Others']
        }.get(resource_type.lower(), [])

        if not resource_tables:
            raise ValueError("Invalid resource type. Use 'labour', 'materials', 'machines' or 'all'")

        results = []
        search_pattern = search_value if exact_match else f'%{search_value}%'

        for table in resource_tables:
            query = f"""
                SELECT 
                    fit.root_id,
                    t.{search_column} as matched_value,
                    tsd.value as title
                FROM {table} t
                JOIN First_Inner_Table fit ON t.first_inner_id = fit.id
                JOIN Title_Section ts ON fit.root_id = ts.root_id
                JOIN Title_Section_Data tsd ON ts.id = tsd.title_section_id AND tsd.key = 'Title'
                WHERE t.{search_column} LIKE ?
            """
            matches = self._execute_query(query, (search_pattern,))
            results.extend([{
                'table_type': table,
                'root_id': row['root_id'],
                'matched_column': search_column,
                'matched_value': row['matched_value'],
                'title': row['title']
            } for row in matches])

        return results

    def get_table_data(
            self,
            root_id: int,
            data_type: str = 'all'  # 'all', 'basic', 'overhead', 'rates', 'references'
    ) -> Dict[str, Any]:
        """
        Get comprehensive data for a specific table
        Args:
            root_id: The root ID of the table
            data_type: Type of data to retrieve
        Returns:
            Dictionary with requested data
        """
        result = {}

        # Always include basic info
        if data_type in ['all', 'basic']:
            result['basic'] = {
                'title': self._get_title_section(root_id).get('Title', [''])[0],
                'references': self._get_references(root_id)
            }

        # Include first inner table data
        if data_type in ['all', 'basic', 'rates']:
            first_inner = self._get_first_inner_table(root_id)
            result['resources'] = {
                'manpower': first_inner.get('Manpower', []),
                'materials': first_inner.get('Materials', []),
                'machines': first_inner.get('Machines', [])
            }

        # Include second inner table data
        if data_type in ['all', 'overhead']:
            second_inner = self._get_second_inner_table(root_id)
            result['costs'] = {
                'overhead': second_inner.get('Overhead', []),
                'rates': second_inner.get('Unit Rate', [])
            }

        return result
    def _get_title_section(self, root_id: int) -> Dict[str, List]:
        """Get title section data for a specific root ID"""
        query = """
            SELECT key, value FROM Title_Section_Data
            WHERE title_section_id = ?
        """
        data = {}
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, (root_id,))
            for key, value in cursor.fetchall():
                if key not in data:
                    data[key] = []
                data[key].append(value)
        except sqlite3.Error as e:
            print(f"Error getting title section: {e}")
        return data

    def _get_references(self, root_id: int) -> Dict[str, List]:
        """Get references data for a specific root ID"""
        query = """
            SELECT key, value FROM Reference_Data_Items
            WHERE reference_data_id = ?
        """
        data = {}
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, (root_id,))
            for key, value in cursor.fetchall():
                if key not in data:
                    data[key] = []
                # Try to convert numeric values
                try:
                    data[key].append(float(value) if '.' in value else int(value))
                except ValueError:
                    data[key].append(value)
        except sqlite3.Error as e:
            print(f"Error getting references: {e}")
        return data

    def _get_first_inner_table(self, root_id: int) -> Dict[str, List]:
        """Get first inner table data for a specific root ID"""
        data = {
            'Title': self._get_inner_title(root_id),
            'Manpower': self._get_inner_data('Manpower', root_id),
            'Materials': self._get_inner_data('Materials', root_id),
            'Machines': self._get_inner_data('Machines', root_id),
            'Others': self._get_inner_data('Others', root_id)
        }
        return data

    def _get_second_inner_table(self, root_id: int) -> Dict[str, List]:
        """Get second inner table data for a specific root ID"""
        query = """
            SELECT key, value FROM Second_Inner_Data
            WHERE second_inner_id = ?
        """
        data = {}
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, (root_id,))
            for key, value in cursor.fetchall():
                if key not in data:
                    data[key] = []
                # Try to convert numeric values
                try:
                    data[key].append(float(value) if '.' in value else int(value))
                except ValueError:
                    data[key].append(value)
        except sqlite3.Error as e:
            print(f"Error getting second inner table: {e}")
        return data

    def _get_inner_title(self, root_id: int) -> List[List[str]]:
        """Get inner table title rows"""
        query = """
            SELECT value FROM First_Inner_Title
            WHERE first_inner_id = ?
            ORDER BY id
        """
        titles = []
        current_row = []
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, (root_id,))
            for (value,) in cursor.fetchall():
                current_row.append(value)
                if len(current_row) >= 7:  # Assuming 7 columns per row
                    titles.append(current_row)
                    current_row = []
            if current_row:  # Add any remaining items
                titles.append(current_row)
        except sqlite3.Error as e:
            print(f"Error getting inner title: {e}")
        return titles

    def _get_inner_data(self, table: str, root_id: int) -> List[List]:
        """Get data from specific inner table"""
        query = f"""
            SELECT * FROM {table}
            WHERE first_inner_id = ?
        """
        try:
            cursor = self.conn.cursor()
            cursor.execute(query, (root_id,))
            return [list(row) for row in cursor.fetchall()]
        except sqlite3.Error as e:
            print(f"Error getting {table} data: {e}")
            return []

    def find_tables(
            self,
            search_criteria: Dict[str, Any],
            max_results: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """
        Find tables matching multiple search criteria
        Args:
            search_criteria: {
                'search_type': 'title'|'resource',
                # For title search:
                'title_pattern': str,
                'exact_match': bool (default False),
                # For resource search:
                'resource_type': 'labour'|'materials'|'machines'|'all' (required for resource search),
                'search_column': str (required for resource search),
                'search_value': str (required for resource search)
            }
            max_results: optional limit on number of results
        Returns:
            List of matching tables with basic info
        """
        search_type = search_criteria.get('search_type')

        if search_type == 'title':
            # Handle title search
            if 'title_pattern' not in search_criteria:
                raise ValueError("title_pattern is required for title search")
            return self.find_tables_by_title(
                title_pattern=search_criteria['title_pattern'],
                exact_match=search_criteria.get('exact_match', False),
                max_results=max_results
            )
        elif search_type == 'resource':
            # Handle resource search
            required = ['search_column', 'search_value']  # Removed resource_type from required
            for field in required:
                if field not in search_criteria:
                    raise ValueError(f"{field} is required for resource search")

            # Default to searching all resource types if not specified
            resource_type = search_criteria.get('resource_type', 'all')

            return self.search_resources(
                resource_type=resource_type,
                search_column=search_criteria['search_column'],
                search_value=search_criteria['search_value'],
                exact_match=search_criteria.get('exact_match', False)
            )
        # elif search_type == 'resource':
        #     # Handle resource search
        #     required = ['resource_type', 'search_column', 'search_value']
        #     for field in required:
        #         if field not in search_criteria:
        #             raise ValueError(f"{field} is required for resource search")
        #     return self.search_resources(
        #         resource_type=search_criteria['resource_type'],
        #         search_column=search_criteria['search_column'],
        #         search_value=search_criteria['search_value'],
        #         exact_match=search_criteria.get('exact_match', False)
        #     )
        else:
            raise ValueError("Invalid search_type. Use 'title' or 'resource'")

    def extract_table_contents(
            self,
            table_ids: List[int],
            content_types: List[str] = ['basic', 'resources', 'costs']
    ) -> Dict[int, Dict[str, Any]]:
        """
        Extract specified content from multiple tables
        Args:
            table_ids: List of root_ids to extract
            content_types: Types of data to extract
        Returns:
            Dictionary with root_id as key and extracted data as value
        """
        results = {}

        for table_id in table_ids:
            table_data = {}

            if 'basic' in content_types:
                table_data['basic'] = {
                    'title': self._get_title_section(table_id).get('Title', [''])[0],
                    'references': self._get_references(table_id)
                }

            if 'resources' in content_types:
                first_inner = self._get_first_inner_table(table_id)
                table_data['resources'] = {
                    'manpower': first_inner.get('Manpower', []),
                    'materials': first_inner.get('Materials', []),
                    'machines': first_inner.get('Machines', [])
                }

            if 'costs' in content_types:
                second_inner = self._get_second_inner_table(table_id)
                table_data['costs'] = {
                    'overhead': second_inner.get('Overhead', []),
                    'rates': second_inner.get('Unit Rate', [])
                }

            results[table_id] = table_data

        return results

    def find_tables_by_title(
            self,
            title_pattern: str,
            exact_match: bool = False,
            max_results: Optional[int] = None
    ) -> List[Dict[str, Any]]:
        """
        Find tables matching title pattern
        Args:
            title_pattern: String to search in titles
            exact_match: Whether to require exact match
            max_results: Optional limit on number of results
        Returns:
            List of matching tables with basic info
        """
        query = """
            SELECT 
                tsd.title_section_id as root_id,
                tsd.value as title
            FROM Title_Section_Data tsd
            WHERE tsd.key = 'Title'
        """

        if exact_match:
            query += " AND tsd.value = ?"
            params = (title_pattern,)
        else:
            query += " AND tsd.value LIKE ?"
            params = (f'%{title_pattern}%',)

        if max_results is not None:
            query += f" LIMIT {max_results}"

        return self._execute_query(query, params)

# Usage Examples
databasepath = "DUDBC_RateAnalysis.db"
# DUDBC_RateWriter(databasepath)


def DUDBC_Extractor(searchmode, searchvalue, resourcetype = ""):
    search_criteria= {}
    if searchmode == "title":
        search_criteria = {
            # # Title search example:
            'search_type': searchmode,
            'title_pattern': searchvalue,
            'exact_match': False
        }

    if searchmode == "resource":
        search_criteria = {
            'search_type': searchmode,
            'search_column': 'level_type',
            'resource_type': resourcetype,  # Optional - will search all types if omitted
            'search_value': searchvalue

        }

    db_path = "DUDBC_RateAnalysis.db"
    db_path = resourece_path(db_path)

    with SuperDataExtractor(db_path) as extractor:
        # Perform the search
        matched_tables = extractor.find_tables(search_criteria)
        all_tables_data = {}

        if not matched_tables:
            print("No matching tables found")
        else:
            # Dictionary to store all tables in the requested format
            all_tables_data = {}

            for table in matched_tables:
                table_id = table['root_id']

                # Get all data for this table
                full_data = {
                    'Title_Section': {},
                    'References': {},
                    'First Inner table': {
                        'Title': [],
                        'Manpower': [],
                        'Materials': [],
                        'Machines': [],
                        'Others': []
                    },
                    'Second Inner table': {}
                }

                # 1. Get Title Section
                title_data = extractor._get_title_section(table_id)
                full_data['Title_Section'] = title_data

                # 2. Get References
                ref_data = extractor._get_references(table_id)
                full_data['References'] = ref_data

                # 3. Get First Inner Table
                first_inner = extractor._get_first_inner_table(table_id)
                def indexContente_splitter(list, title):
                    Contents = []
                    Indexes = []
                    for table in list:
                        Contents.append(table[2:])
                        Indexes.append(table[0:2])

                    full_data['First Inner table'][title] = Contents

                indexContente_splitter(first_inner['Manpower'], "Manpower")
                indexContente_splitter(first_inner['Materials'], "Materials")
                indexContente_splitter(first_inner['Machines'], "Machines")
                full_data['First Inner table']['Others'] = first_inner['Others']



                # 4. Get Second Inner Table
                second_inner = extractor._get_second_inner_table(table_id)
                full_data['Second Inner table'] = second_inner

                # Add to our complete collection
                all_tables_data[table_id] = full_data

            # Print the first table's data in the exact requested format
            # print(all_tables_data)
            TableKeys = list(all_tables_data.keys())
            # for table_id in TableKeys:
            #     print(f"{table_id}: {all_tables_data[table_id]}")
                ### printsample = 187: {'Title_Section': {'Title': ['नयाँ सर्फेसमा ह्वाईटवास दुईकोट गर्ने काम (सिलिंगमा)'], 'Note': ['दर विश्लेषणको लागि १०० व.मी. लिइएको']}, 'References': {'Table No': [192], 'Reference': ['J 1']}, 'First Inner table': {'Title': [], 'Manpower': [['श्रमिक', 'क) सिपालु', 1.875, 'जवान', 1190.0, 2231.25, None], ['श्रमिक', 'ख) ज्यामी', 1.375, 'जवान', 863.0, 1186.6200000000001, 3417.87]], 'Materials': [['निर्माण सामग्री', 'सेतो चुना', 22.0, 'के.जी.', 21.25, 467.5, None], ['निर्माण सामग्री', 'गम आदि', 0.88, 'के.जी.', 290.7, 255.81, 723.31]], 'Machines': [], 'Others': []}, 'Second Inner table': {'Original Rate': [4141.18], 'Overhead': [621.17], 'Total Rate': [4762.35], 'Unit Rate': [47.62], 'Others': ['दर प्रति व.मी.को', 4762.35, 100]}}


            with open('all_tables_data.json', 'w', encoding='utf-8') as f:
                json.dump(all_tables_data, f, indent=2, ensure_ascii=False)
        return all_tables_data



# DUDBC_Extractor("title", " सिमेन्ट पेन्ट", resourcetype = "")
# ####'labour', 'materials', 'machines' or 'all'
# DUDBC_Extractor("resource", "डिस्टेम्पर पाउडर", resourcetype = "all")

#Results
# 194: {'Title_Section': {'Title': ['वाटर प्रुफ सिमेन्ट पेन्ट एक कोट लगाउने काम ।'], 'Note': ['दर विश्लेषणको लागि १०० व.मी. लिइएको']}, 'References': {'Table No': [199], 'Reference': ['J 4']}, 'First Inner table': {'Title': [], 'Manpower': [['श्रमिक', 'क) सिपालु', 1.7, 'जवान', 1190.0, 2023.0, None], ['श्रमिक', 'ख) ज्यामी', 1.7, 'जवान', 863.0, 1467.1000000000001, 3490.1000000000004]], 'Materials': [['निर्माण सामग्री', 'सिमेन्ट पेन्ट', 30.0, 'के.जी.', 42.84, 1285.2, 1285.2], ['निर्माण सामग्री', None, None, None, None, None, None]], 'Machines': [], 'Others': []}, 'Second Inner table': {'Original Rate': [4775.3], 'Overhead': [716.29], 'Total Rate': [5491.59], 'Unit Rate': [54.91], 'Others': ['दर प्रति व.मी.को', 5491.59, 100]}}
# 195: {'Title_Section': {'Title': ['वाटर प्रुफ सिमेन्ट पेन्ट दुई कोट लगाउने काम ।'], 'Note': ['दर विश्लेषणको लागि १०० व.मी. लिइएको']}, 'References': {'Table No': [200], 'Reference': ['J 4']}, 'First Inner table': {'Title': [], 'Manpower': [['श्रमिक', 'क) सिपालु', 5.0, 'जवान', 1190.0, 5950.0, None], ['श्रमिक', 'ख) ज्यामी', 5.0, 'जवान', 863.0, 4315.0, 10265.0]], 'Materials': [['निर्माण सामग्री', 'सिमेन्ट पेन्ट', 48.5, 'के.जी.', 42.84, 2077.7400000000002, 2077.7400000000002], ['निर्माण सामग्री', None, None, None, None, None, None]], 'Machines': [], 'Others': []}, 'Second Inner table': {'Original Rate': [12342.74], 'Overhead': [1851.41], 'Total Rate': [14194.15], 'Unit Rate': [141.94], 'Others': ['दर प्रति व.मी.को', 14194.15, 100]}}
# 192: {'Title_Section': {'Title': ['डिस्टेम्पर लगाउने काम एक कोट ।'], 'Note': ['दर विश्लेषणको लागि १०० व.मी. लिइएको']}, 'References': {'Table No': [197], 'Reference': ['J 3']}, 'First Inner table': {'Title': [], 'Manpower': [['श्रमिक', 'क) सिपालु', 4.0, 'जवान', 1190.0, 4760.0, None], ['श्रमिक', 'ख) ज्यामी', 4.0, 'जवान', 863.0, 3452.0, 8212.0]], 'Materials': [['निर्माण सामग्री', 'अस्तर', 8.0, 'ली.', 397.8, 3182.4, None], ['निर्माण सामग्री', 'डिस्टेम्पर पाउडर', 6.5, 'के.जी.', 163.2, 1060.8, 4243.2]], 'Machines': [], 'Others': []}, 'Second Inner table': {'Original Rate': [12455.2], 'Overhead': [1868.28], 'Total Rate': [14323.480000000001], 'Unit Rate': [143.23], 'Others': ['दर प्रति व.मी.को', 14323.480000000001, 100]}}
# 193: {'Title_Section': {'Title': ['डिस्टेम्पर लगाउने काम दुई कोट ।'], 'Note': ['दर विश्लेषणको लागि १०० व.मी. लिइएको']}, 'References': {'Table No': [198], 'Reference': ['J 3']}, 'First Inner table': {'Title': [], 'Manpower': [['श्रमिक', 'क) सिपालु', 5.8, 'जवान', 1190.0, 6902.0, None], ['श्रमिक', 'ख) ज्यामी', 5.8, 'जवान', 863.0, 5005.400000000001, 11907.400000000001]], 'Materials': [['निर्माण सामग्री', 'अस्तर', 8.0, 'ली.', 397.8, 3182.4, None], ['निर्माण सामग्री', 'डिस्टेम्पर पाउडर', 11.5, 'के.जी.', 163.2, 1876.8, 5059.2]], 'Machines': [], 'Others': []}, 'Second Inner table': {'Original Rate': [16966.600000000002], 'Overhead': [2544.9900000000002], 'Total Rate': [19511.590000000004], 'Unit Rate': [195.11], 'Others': ['दर प्रति व.मी.को', 19511.590000000004, 100]}}








#DatabaseWrite

class GUIDatabase:
    #For data modificaiton (ie adding of any unique key and values) then perfrom the following:
    #1. initialize from the table creation
    #2. Enforce the code for writing to that key
    def __init__(self, db_name="estimation.db", init_db=True):
        if init_db:
            db_name  = resourece_path(db_name)
            if os.path.exists(db_name):
                os.remove(db_name)

            # Create a fresh database
            self.conn = sqlite3.connect(db_name)
            self.cursor = self.conn.cursor()
            self.init_db()
            self.conn = sqlite3.connect(db_name)
            self.cursor = self.conn.cursor()
            self.init_db()

    def init_db(self):
        # General Info table
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS General_Info (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            office TEXT,
            projectname TEXT,
            officeCode TEXT,
            projectlocation TEXT,
            projectcompletiontime TEXT,
            fiscalyear TEXT,
            budgetsubheadingno TEXT
        )
        """)

        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS quantity_estimation (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            EstimationPartSection_root TEXT,
            estimation_Section_title TEXT,
            items_section_Title TEXT,
            dropdown TEXT,
            search_keyword_input TEXT,
            search_button TEXT,
            dynamic_saerchResults_container TEXT,
            item_number TEXT,
            item_description TEXT,
            unit TEXT,
            rate REAL,
            numbers REAL,
            length REAL,
            breadth REAL,
            height REAL,
            quantity REAL,
            remarks TEXT,
            calc_info TEXT,
            quantity_factor REAL,
            Item_cost REAL
        )
        """)



        # Search Results (applied items)
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS Rate_Analysis (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_number TEXT NOT NULL,
            norms_ref TEXT,
            section TEXT,
            attribute TEXT,
            value TEXT,
            resource_type TEXT,
            category TEXT,
            quantity REAL,
            unit TEXT,
            rate REAL,
            amount REAL,
            amount_perHeading REAL
        )
        """)

        #Create table for holding the primary factors of the data
        self.cursor.execute("""
        CREATE TABLE IF NOT EXISTS Primary_keys (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contingency REAL,
            vat REAL,
            physical_contingency REAL,
            priceadjustment_contingency REAL
        )
        """)
        self.conn.commit()
    def save_GenInfo_QEstimation(self, ObjectsCache):
        def get_text(obj):
            try:
                return str(obj.text)
            except Exception:
                return str(obj)  # fallback if no .text attr

        def safe_float(val):
            try:
                return float(val)
            except (ValueError, TypeError):
                return val

        def derive_Number(val, default=0.0):
            try:
                val = val.split(":")
                val = val[-1]
                return float(val)
            except (ValueError, TypeError):
                return default


        # --- Save General Information ---
        general_info = ObjectsCache["Estimation_Data"]["General_Information"]

        general_values = (
            get_text(general_info["office"]),
            get_text(general_info["projectname"]),
            get_text(general_info["officeCode"]),
            get_text(general_info["projectlocation"]),
            get_text(general_info["projectcompletiontime"]),
            get_text(general_info["fiscalyear"]),
            get_text(general_info["budgetsubheadingno"]),
        )

        # Clear existing rows and insert fresh one
        self.cursor.execute("DELETE FROM General_Info")
        self.cursor.execute("""
            INSERT INTO General_Info (
                office, projectname, officeCode, projectlocation,
                projectcompletiontime, fiscalyear, budgetsubheadingno
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, general_values)

        # --- Save Quantity Estimation ---
        sections = ObjectsCache["Estimation_Data"]["Estimation_Sections"]

        for _, section in sections.items():
            item_number = get_text(section["item_number"])

            values = (
                str(section["EstimationPartSection_root"]),
                get_text(section["estimation_Section_title"]),
                get_text(section["items_section_Title"]),
                get_text(section["dropdown"]),
                get_text(section["search_keyword_input"]),
                get_text(section["search_button"]),
                get_text(section["dynamic_saerchResults_container"]),
                item_number,
                get_text(section["item_description"]),
                get_text(section["unit"]),
                safe_float(get_text(section["rate"])),
                safe_float(get_text(section["numbers"])),
                safe_float(get_text(section["length"])),
                safe_float(get_text(section["breadth"])),
                safe_float(get_text(section["height"])),
                safe_float(get_text(section["quantity"])),
                get_text(section["remarks"]),
                str(section["calc_info"]),
                derive_Number(get_text(section["quantity_factor"])),
                derive_Number(get_text(section["Item_cost"])),
            )
            # print(values)

            # Check if row exists with same item_number
            self.cursor.execute("SELECT id FROM quantity_estimation WHERE item_number = ?", (item_number,))
            existing = self.cursor.fetchone()

            if existing:
                # Delete old row
                self.cursor.execute("DELETE FROM quantity_estimation WHERE item_number = ?", (item_number,))

            # Insert new/updated row
            self.cursor.execute("""
                INSERT INTO quantity_estimation (
                    EstimationPartSection_root, estimation_Section_title,
                    items_section_Title, dropdown, search_keyword_input, search_button,
                    dynamic_saerchResults_container, item_number, item_description,
                    unit, rate, numbers, length, breadth, height, quantity,
                    remarks, calc_info, quantity_factor, Item_cost
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, values)

        self.conn.commit()

    def save_PrimaryKeys(self, valuesDict):
        def safe_float(val, default=0.0):
            try:
                return float(val)
            except (ValueError, TypeError):
                return default



        # --- Save General Information ---
        primary_values = (safe_float(valuesDict["contingency"]),safe_float(valuesDict["vat"]),safe_float(valuesDict["physical_contingency"]),safe_float(valuesDict["price_adjustment"]), )

        # Clear existing rows and insert fresh one
        self.cursor.execute("DELETE FROM Primary_keys")
        self.cursor.execute("""
            INSERT INTO Primary_keys (
                contingency, vat, physical_contingency, priceadjustment_contingency
            ) VALUES (?, ?, ?, ?)
        """, primary_values)


        self.conn.commit()

    def close(self):
        self.conn.close()



#_________________-Check below

    def save_appliedRateAnalysis(self, item_number, appliedRateData):
        db_name = resourece_path("estimation.db")
        conn = sqlite3.connect(db_name)
        c = conn.cursor()

        # First remove all rows with this item_number
        c.execute("DELETE FROM Rate_Analysis WHERE item_number = ?", (item_number,))

        # ---- Title_Section ----
        for k, v in appliedRateData.get("Title_Section", {}).items():
            for entry in v:
                c.execute("""
                    INSERT INTO Rate_Analysis (item_number, norms_ref, section, attribute, value)
                    VALUES (?, ?, ?, ?, ?)
                """, (item_number, appliedRateData.get("NormsDBRef", ""), "Title_Section", k, str(entry)))

        # ---- References ----
        for k, v in appliedRateData.get("References", {}).items():
            for entry in v:
                c.execute("""
                    INSERT INTO Rate_Analysis (item_number, norms_ref, section, attribute, value)
                    VALUES (?, ?, ?, ?, ?)
                """, (item_number, appliedRateData.get("NormsDBRef", ""), "References", k, str(entry)))

        # ---- First Inner Table ----
        inner = appliedRateData.get("First Inner table", {})

        def insert_resource(rows, resource_type):
            for row in rows:
                c.execute("""
                    INSERT INTO Rate_Analysis (
                        item_number, norms_ref, section, attribute,
                        resource_type, category, quantity, unit, rate, amount, amount_perHeading
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    item_number, appliedRateData.get("NormsDBRef", ""),
                    "First Inner table", resource_type,
                    row[0], row[1], row[2], row[3], row[4], row[5],
                    row[6] if len(row) > 6 else None  # 7th item stored in extra
                ))

        insert_resource(inner.get("Manpower", []), "Manpower")
        insert_resource(inner.get("Materials", []), "Materials")
        insert_resource(inner.get("Machines", []), "Machines")

        # ---- Second Inner Table ----
        for k, v in appliedRateData.get("Second Inner table", {}).items():
            for entry in v:
                c.execute("""
                    INSERT INTO Rate_Analysis (item_number, norms_ref, section, attribute, value)
                    VALUES (?, ?, ?, ?, ?)
                """, (item_number, appliedRateData.get("NormsDBRef", ""), "Second Inner table", k, str(entry)))

        conn.commit()
    def load_appliedRateAnalysis(self, item_number=None, norms_ref=None):
        """
        Extract data from Rate_Analysis table and reconstruct appliedRateData
        in the exact format it was saved.

        Parameters:
            item_number (str): specific item_number to filter.
            norms_ref (str/int): specific NormsDBRef to filter.

        Returns:
            appliedRateData (dict)
        """
        conn = sqlite3.connect("estimation.db")
        c = conn.cursor()

        # Initialize the structure
        appliedRateData = {
            "NormsDBRef": "",
            "Title_Section": {},
            "References": {},
            "First Inner table": {
                "Title": [["स्रोत साधन", "तह/किसिम", "परिमाण", "एकाई", "दर प्रति एकाई", "रकम",
                           "प्रत्येक स्रोत साधनको जम्मा"]],
                "Manpower": [],
                "Materials": [],
                "Machines": [],
                "Others": []
            },
            "Second Inner table": {}
        }

        # Build query based on filters
        conn = sqlite3.connect("estimation.db")
        c = conn.cursor()

        # Base query
        query = "SELECT * FROM Rate_Analysis WHERE 1=1"
        params = []

        # Filter by item_number if provided
        if item_number is not None:
            query += " AND item_number=?"
            params.append(item_number)

        # Filter by norms_ref if provided
        if norms_ref is not None:
            query += " AND norms_ref=?"
            params.append(norms_ref)

        # Execute query
        c.execute(query, params)
        rows = c.fetchall()

        if not rows:
            conn.close()
            return rows,  appliedRateData

        # Get column names
        col_names = [desc[0] for desc in c.description]

        # Fetch NormsDBRef from first non-empty row
        for row in rows:
            idx = col_names.index("norms_ref")
            if row[idx]:
                appliedRateData["NormsDBRef"] = row[idx]
                break

        # Process each row
        for row in rows:
            section = row[col_names.index("section")]
            attribute = row[col_names.index("attribute")]

            if section == "Title_Section":
                value = row[col_names.index("value")]
                if attribute not in appliedRateData["Title_Section"]:
                    appliedRateData["Title_Section"][attribute] = []
                appliedRateData["Title_Section"][attribute].append(value)

            elif section == "References":
                value = row[col_names.index("value")]
                if attribute not in appliedRateData["References"]:
                    appliedRateData["References"][attribute] = []
                appliedRateData["References"][attribute].append(value)

            elif section == "First Inner table":
                resource_type = row[col_names.index("attribute")]
                entry = [
                    row[col_names.index("resource_type")],
                    row[col_names.index("category")],
                    row[col_names.index("quantity")],
                    row[col_names.index("unit")],
                    row[col_names.index("rate")],
                    row[col_names.index("amount")],
                    row[col_names.index("amount_perHeading")]
                ]
                if resource_type in ["Manpower", "Materials", "Machines"]:
                    appliedRateData["First Inner table"][resource_type].append(entry)
                else:
                    appliedRateData["First Inner table"]["Others"].append(entry)

            elif section == "Second Inner table":
                value = row[col_names.index("value")]
                if attribute not in appliedRateData["Second Inner table"]:
                    appliedRateData["Second Inner table"][attribute] = []
                appliedRateData["Second Inner table"][attribute].append(value)

        conn.close()
        return rows, appliedRateData

    def ResetRate_QunatityEstimation(self):
        conn = sqlite3.connect("estimation.db")
        c = conn.cursor()

        #Deletes the whole quantity_estimation table
        c.execute("DELETE FROM quantity_estimation")  # remove all rows
        c.execute("DELETE FROM sqlite_sequence WHERE name='quantity_estimation'")  # reset autoincrement
        conn.commit()

        #Deletes the whole quantity_estimation table
        c.execute("DELETE FROM Rate_Analysis")  # remove all rows
        c.execute("DELETE FROM sqlite_sequence WHERE name='Rate_Analysis'")  # reset autoincrement
        conn.commit()

        conn.close()


    def delete_SubItemData_(self, item_number):
        conn = sqlite3.connect("estimation.db")
        c = conn.cursor()

        c.execute("SELECT id FROM quantity_estimation WHERE item_number = ?", (item_number,))
        existing = c.fetchone()
        # print("_____________________________________________________________")
        # print(existing, "existing", item_number)

        if existing:
            c.execute("DELETE FROM quantity_estimation WHERE item_number = ?", (item_number,))
            conn.commit()

        c.execute("SELECT id FROM quantity_estimation WHERE item_number = ?", (item_number,))
        existing = c.fetchone()
        # print(existing, "existing")
        conn.close()






class DB_Output:
    def __init__(self, db_name="estimation.db"):
        self.db_GUI = GUIDatabase(init_db=False)
        self.conn = sqlite3.connect(db_name)
        self.cursor = self.conn.cursor()
        self.excel_name = "output.xlsx"
        self.base_dir = os.path.dirname(os.path.abspath(self.excel_name))  # same folder as Excel file
        self.pdf_path = os.path.join(self.base_dir, "CombinedBinder.pdf")  # full path

        self.sheetnames = ["Cover" , "Quantity Estimation", "Summary", "Abstract of Cost", "BOQ", "Rate Analysis" ]
        self.sheetnamesTITLEMerger_Range = {"Quantity Estimation": ["A", "I", "H", 9],"Summary": ["A", "E", "D", 9],"Abstract of Cost": ["A", "G", "F", 9],"BOQ": ["A", "G", "F", 9],"Cover": ["A", "I", "H", 9],"Rate Analysis": ["A", "H", "G", 7]}   #For every sheets they are ["start column", "end column", "FY writing column", Freezing row]
        self.headers = [            "S.No.", "Description of Works", "Unit", "No.",            "Length (m)", "Breadth (m)", "Height (m)", "Quantity", "Remarks"        ]
        self.AOCheaders = [   "S.N.", "Description of Works", "Unit", "Quantity", "Rate", "Amount" , "Remarks"]
        self.Summaryheaders = [   "S.N.", "Description of Works", "Quantity","Unit", "Remarks"]
        self.BOQheaders = [   "S.N.", "Description of Works","Unit","Quantity",  "Bidders Rate NRs (In figure)", "Bidders Rate NRs (In words)", "Amount (In figure)"]


        self.qEstDBtitles = [ 'EstimationPartSection_root', 'estimation_Section_title', 'items_section_Title', 'dropdown', 'search_keyword_input', 'search_button', 'dynamic_saerchResults_container', 'item_number', 'item_description', 'unit', 'rate', 'numbers', 'length', 'breadth', 'height', 'quantity', 'remarks', 'calc_info', 'quantity_factor', 'Item_cost']

        #Styles
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        self.right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.KMTitleFont = Font(name="Kalimati", size=12, bold=True)
        self.KMbold_font = Font(name="Kalimati", bold=True, size=10)
        self.KMNormal_font = Font(name="Kalimati", bold=False, size=9)
        self.Special_font = Font(name="Times New Roman", bold=True, size=10, underline="single")
        self.TNRbold_font = Font(name="Times New Roman", bold=True, size=10)
        self.TNRnormalText_font = Font(name="Times New Roman", bold=False, size=10)
        self.thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                                  bottom=Side(style="thin"))
        self.thick_border = Border(left=Side(style="thick"), right=Side(style="thick"), top=Side(style="thick"),
                                  bottom=Side(style="thick"))
        self.bottom_border = Border(bottom=Side(style="thin"))
        self.bottom_border_thick = Border(bottom=Side(style="thick"))

        red_thick_side = Side(style="thick", color="FF0000")  # FF0000 is red in hex
        self.thick_Redborder = Border(            left=red_thick_side,            right=red_thick_side,            top=red_thick_side,            bottom=red_thick_side)


        #Top lines layout
        #AB
        #CD
        self.borderA = Border(left=Side(style="thick"), top=Side(style="thick"),    bottom=Side(style="thin"))
        self.borderB = Border(right=Side(style="thick"), top=Side(style="thick"),    bottom=Side(style="thin"))
        self.borderC = Border(left=Side(style="thick"), top=Side(style="thin"),    bottom=Side(style="thick"))
        self.borderD = Border(right=Side(style="thick"), top=Side(style="thin"),    bottom=Side(style="thick"))
    def fetch_general_info(self):
        """Fetch all data from General_Info table"""
        self.cursor.execute("""
            SELECT office, projectname, officeCode, projectlocation, 
                   projectcompletiontime, fiscalyear, budgetsubheadingno
            FROM General_Info
        """)
        return self.cursor.fetchall()
    def fetch_PrimaryKeys(self):
        """Fetch all data from General_Info table"""
        self.cursor.execute("""
            SELECT contingency ,        vat ,        physical_contingency ,        priceadjustment_contingency 
            FROM Primary_keys
        """)
        return self.cursor.fetchall()

    def fetch_quantityEstimation(self):
        """Fetch all data from quantity_estimation table"""
        self.cursor.execute(f"PRAGMA table_info({'quantity_estimation'})")
        titles = [col[1] for col in self.cursor.fetchall()]

        self.cursor.execute("""            SELECT EstimationPartSection_root, estimation_Section_title,            items_section_Title, dropdown, search_keyword_input, search_button,            dynamic_saerchResults_container, item_number, item_description,            unit, rate, numbers, length, breadth, height, quantity,            remarks, calc_info, quantity_factor, Item_cost            FROM quantity_estimation        """)
        rows = self.cursor.fetchall()
        return titles, rows

    def dataTrimming_(self, qEstRows):
        trimmedData = []
        trinFor = ['item_number', 'remarks']  # [start data, end daata]
        startindex, endindex = self.qEstDBtitles.index(trinFor[0]), self.qEstDBtitles.index(trinFor[1])
        # The index extracted are the exact index(ie natural no numbering) as id from the database are with first row id and isnot returned to the call in this code
        for line in qEstRows:
            linedata = line[startindex: endindex + 1]
            trimmedData.append(linedata)

        itemNo_List = set()
        for line in qEstRows:
            item_nos = line[startindex].split(".")
            item_no = ".".join(item_nos[0:2])
            itemNo_List.add(item_no)

        return sorted(itemNo_List, key=lambda x: [int(i) for i in x.split('.')]), trimmedData

    def dataCategorization_(self, itemNo_List, qEstRows):
        itemsNo_index = self.qEstDBtitles.index("item_number")
        categorizedData = {}

        # Initialize the dictionary
        for item in itemNo_List:
            categorizedData[item] = []

        for linerow in qEstRows:
            item_no = ".".join(linerow[itemsNo_index].split(".")[0:2])
            if item_no in itemNo_List:
                categorizedData[item_no].append(linerow)

        return itemsNo_index, categorizedData

    def dataBulking(self, itemNo_List, categorizedData):
        trimmedData = []
        trinFor = ['item_number', 'remarks']  # [start data, end daata]
        startindex, endindex = self.qEstDBtitles.index(trinFor[0]), self.qEstDBtitles.index(trinFor[1])
        # The index extracted are the exact index(ie natural no numbering) as id from the database are with first row id and isnot returned to the call in this code
        bulkedData = {}

        for item in itemNo_List:
            bulkedData[item] = []

        # Addition of preceeding and succeeding row on certain bulks
        for item in itemNo_List:
            datas = categorizedData[item]
            itemstitle_index = self.qEstDBtitles.index("items_section_Title")
            itemTitlevalue = datas[0][itemstitle_index]
            datanos = len(categorizedData[item][0])
            descriptiontitle_index = self.qEstDBtitles.index("item_description")

            preceedinglist = [None] * datanos
            preceedinglist[descriptiontitle_index] = itemTitlevalue

            succeedingList = [None] * datanos
            succeedingList[descriptiontitle_index] = "Sub total"

            itemsquantity_index = self.qEstDBtitles.index("quantity")
            subtotalvalue_index = self.qEstDBtitles.index("quantity_factor")
            subtotalvalue = datas[0][subtotalvalue_index]
            succeedingList[itemsquantity_index] = subtotalvalue

            bulkedData[item].append(preceedinglist)
            for data in datas:
                bulkedData[item].append(data)
            bulkedData[item].append(succeedingList)
        return bulkedData

    def data_segregation(self, bulkedData, needed_keys):
        segregated_dict = {}
        for item_key, rows in bulkedData.items():
            item_list = []
            for row in rows:
                if isinstance(row, (list, tuple)):
                    # full row dict (all fields)
                    full_dict = dict(zip(self.qEstDBtitles, row))
                    # extract only needed fields but keep in original full row as well
                    filtered_dict = [full_dict.get(k) for k in needed_keys]
                    item_list.append(filtered_dict)
            segregated_dict[item_key] = item_list

        return segregated_dict

    def ProcessedData_Extractor(self):
        needed_keys = ['item_number', 'item_description', 'unit', 'numbers', 'length', 'breadth', 'height', 'quantity',
                       'remarks']

        dbTitles, qEstRows = self.fetch_quantityEstimation()
        itemNo_List, trimmedData = self.dataTrimming_(qEstRows)
        itemsNo_index, categorizedData = self.dataCategorization_(itemNo_List, qEstRows)
        bulkedData = self.dataBulking(itemNo_List, categorizedData)
        segregated_dict = self.data_segregation(bulkedData, needed_keys)

        return itemNo_List, trimmedData, categorizedData,bulkedData, segregated_dict

#
    def SheetsTitle_Writing(self):
        data = self.fetch_general_info()
        if not data:
            print("No data found in General_Info")
            return
        record = self.fetch_general_info()[0]

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Create 5 sheets
        for sheet in range(0, 6):
            ws = wb.create_sheet(title=f"{self.sheetnames[sheet]}")



            row = 1
            office_full, projectname, officeCode, projectlocation, completion_time, fiscalyear, budgetsubheadingno = record
            offices = office_full.split('/')  # Split the office hierarchy

            # Write office hierarchy
            mergerrows = f"{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][0]}{row}:{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][1]}{row}"
            ws.merge_cells(mergerrows)
            cell = ws.cell(row=row, column=1)
            cell.value = f"नेपाल सरकार"           #Insert the gui entry to add the location of the office
            cell.font = self.KMTitleFont
            cell.alignment = self.center_align
            row += 1

            for office in offices:
                mergerrows = f"{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][0]}{row}:{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][1]}{row}"
                ws.merge_cells(mergerrows)
                cell = ws[f"A{row}"]
                # Assign value and style
                cell.value = office
                cell.font = self.KMTitleFont
                cell.alignment = self.center_align
                # ws[f"A{row}"].fill = grey_fill
                row += 1

            mergerrows = f"{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][0]}{row}:{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][1]}{row}"
            ws.merge_cells(mergerrows)
            cell = ws.cell(row=row, column=1)
            cell.value = f"{projectlocation}"           #Insert the gui entry to add the location of the office
            cell.font = Font(name="Kalimati", bold=False, size=10)
            cell.alignment = self.center_align
            row += 1

            if self.sheetnames[sheet] != "Cover":
                row += 1
                mergerrows = f"{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][0]}{row}:{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][1]}{row}"
                ws.merge_cells(mergerrows)
                cell = ws.cell(row=row, column=1)
                cell.value = f"योजनाको नामः {projectname}"
                cell.font = self.KMbold_font
                row += 1

                ws.merge_cells(f"A{row}:C{row}")
                cell = ws.cell(row=row, column=1)
                cell.value = f"योजना स्थलः {projectlocation}।"
                cell.font = self.KMbold_font

                # Fiscal year
                mergerrows = f"{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][2]}{row}:{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][1]}{row}"
                ws.merge_cells(mergerrows)
                cell = ws[f"{self.sheetnamesTITLEMerger_Range[self.sheetnames[sheet]][2]}{row}"]
                cell.value=f"F.Y.: {fiscalyear}"


        wb.save(self.excel_name)
        print(f"Data exported successfully to {self.excel_name}")
        wb.close()
        return True
    def CoverPage_Writing(self):
        # Load the existing workbook
        record = self.fetch_general_info()[0]

        office_full, projectname, officeCode, projectlocation, completion_time, fiscalyear, budgetsubheadingno = record

        wb = op.load_workbook(self.excel_name)  # 👈 change filename if needed
        if "Cover" in wb.sheetnames:
            ws = wb["Cover"]
        else:
            print("Cover' not found!")
            exit()
        row = ws.max_row + 1       #From here the writing needs to be started

        #Establishing the definite column widths
        column_widths = [8,8,12,8,8,8,8,4,8]
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        sheettitle = "विस्तृत परिमाण र लागत अनुमान"
        mergerrows = f"A5:I6"
        ws.merge_cells(mergerrows)
        cell = ws.cell(row=5, column=1)
        cell.value = f"{sheettitle}"  # Insert the gui entry to add the location of the office
        cell.font = Font(name="Kalimati", bold=False, size=20)
        cell.alignment = self.center_align


        # Define a thick border style           # Define columns for vertical lines (e.g., B, E, and H)
        thick = Side(border_style="thick", color="000000")
        both_border = Border(right=thick, left=thick)
        right_border = Border(right=thick)
        line_columns = ['D', 'E']
        # Define line height (rows)
        for col in line_columns:
            for index, row in enumerate(range(8, 17)):  # Adjust height (rows 2 to 30)
                if index == 0:
                    if col == line_columns[0]:
                        ws[f"{col}{row}"].border = right_border
                        print(row, col)
                elif index == 8:
                    if col == line_columns[0]:

                        ws[f"{col}{row}"].border = right_border
                else:
                    ws[f"{col}{row}"].border = both_border


        mergerrows = f"A18:I18"
        ws.merge_cells(mergerrows)
        cell = ws.cell(row=18, column=1)
        cell.value = f"योजनाको नामः {projectname}"
        cell.font = Font(name="Kalimati", bold=True, size=14)
        cell.alignment = self.center_align

        mergerrows = f"A19:I19"
        ws.merge_cells(mergerrows)
        cell = ws.cell(row=19, column=1)
        cell.value = f"योजना स्थलः {projectlocation}।"
        cell.font = Font(name="Kalimati", bold=True, size=14)
        cell.alignment = self.center_align

        mergerrows = f"A20:I20"
        ws.merge_cells(mergerrows)
        cell = ws.cell(row=20, column=1)
        cell.value = f"आर्थिक वर्ष: {fiscalyear}।"
        cell.font = Font(name="Kalimati", bold=True, size=14)
        cell.alignment = self.center_align



        wb.save(self.excel_name)
        print(f"Data written successfully to Cover in {self.excel_name}")
        wb.close()
        return True

    def QuantityEstSheet_Writing(self):
        # Load the existing workbook

        wb = op.load_workbook(self.excel_name)  # 👈 change filename if needed
        if "Quantity Estimation" in wb.sheetnames:
            ws = wb["Quantity Estimation"]
        else:
            print("Sheet 'Quantity estimation' not found!")
            exit()
        row = ws.max_row + 1       #From here the writing needs to be started

        def Titles_writing(row):
            mergerrows = f"{self.sheetnamesTITLEMerger_Range['Quantity Estimation'][0]}{row}:{self.sheetnamesTITLEMerger_Range['Quantity Estimation'][1]}{row}"
            ws.merge_cells(mergerrows)
            cell = ws.cell(row=row, column=1)
            cell.value = f"Quantity Estimation {12}"
            cell.font = self.Special_font
            cell.alignment = self.center_align
            row += 1

            # Writing for the titles of estimation
            for col, header in enumerate(self.headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.TNRbold_font
                cell.alignment = self.center_align
                cell.border = self.thin_border

            # Optional: Adjust column widths for readability
            column_widths = [5, 25, 5, 4.5, 7, 7, 7, 7, 8]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width
            return row

        def QuantityEstimationWriting(segregated_dict, row):
            # ✅ Write segregated_dict to Excel
            for item_key in sorted(segregated_dict.keys(), key=lambda x: [int(i) for i in x.split('.')]):
                rows_block = segregated_dict[item_key]
                for i, inner_row in enumerate(rows_block):
                    for col, value in enumerate(inner_row, start=1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.TNRnormalText_font

                        # ✅ Alignment rules
                        if col == 2:  # description column
                            cell.alignment = self.left_align
                        else:
                            cell.alignment = self.center_align

                        if i == len(rows_block) - 1:  # last row of block
                            cell.alignment = self.right_align

                        cell.border = self.thin_border
                    row += 1  # move to next row after writing one line


        #___________________________________________________________________Calling and Writing
        row = Titles_writing(row)
        row +=1
        itemNo_List, trimmedData, categorizedData, bulkedData, segregated_dict = self.ProcessedData_Extractor()
        QuantityEstimationWriting(segregated_dict, row)


        wb.save(self.excel_name)
        print(f"Data written successfully to Quantity Estimation in {self.excel_name}")
        wb.close()
        # return   itemNo_List, trimmedData, categorizedData,bulkedData, segregated_dict
        return True

    def AOC_writing(self):
        DescriptionTitle_index = 1  #Zero based indexing
        QuantityTitle_index = 7  #Zero based indexing
        Unit_Bulkindex = self.qEstDBtitles.index("unit")
        Rate_Bulkindex= self.qEstDBtitles.index("rate")
        wb = op.load_workbook(self.excel_name)  # 👈 change filename if needed
        if "Abstract of Cost" in wb.sheetnames:
            ws = wb["Abstract of Cost"]
        else:
            print("Sheet 'Quantity estimation' not found!")
            exit()
        row = ws.max_row + 1       #From here the writing needs to be started


        itemNo_List, trimmedData, categorizedData, bulkedData, segregated_dict = self.ProcessedData_Extractor()
        def Titles_writing(row):
            mergerrows = f"{self.sheetnamesTITLEMerger_Range['Abstract of Cost'][0]}{row}:{self.sheetnamesTITLEMerger_Range['Abstract of Cost'][1]}{row}"
            ws.merge_cells(mergerrows)
            cell = ws.cell(row=row, column=1)
            cell.value = f"Abstract of Cost {12}"
            cell.font = self.Special_font
            cell.alignment = self.center_align
            row += 1

            # Writing for the titles of estimation
            for col, header in enumerate(self.AOCheaders, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.TNRbold_font
                cell.alignment = self.center_align
                cell.border = self.thin_border
                cell.border = self.bottom_border_thick


            # Optional: Adjust column widths for readability
            column_widths = [5, 30, 7, 7, 8, 10, 7]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width
            return row
        def datasplitting(segregated_dict):
            reduced_dict = {}
            for key, rows in segregated_dict.items():
                if rows:  # make sure not empty
                    reduced_dict[key] = [rows[0], rows[-1]]
            return reduced_dict
        def item_unit():
            item_unitDict = {}
            item_rateDict = {}
            for key, value in categorizedData.items():
                item_unitDict[key] = value[0][Unit_Bulkindex]
                item_rateDict[key] = value[0][Rate_Bulkindex]


            return item_unitDict, item_rateDict
        def AOC_Data(reduced_dict):
            AOCList  = []

            for key, value in reduced_dict.items():
                AOCLine = []  ### "S.N.", "Description of Works", "Unit", "Quantity", "Rate", "Amount" , "Remarks"
                quanitity = value[1][QuantityTitle_index]
                rate = item_rateDict[key]

                AOCLine.append(key)
                AOCLine.append(value[0][DescriptionTitle_index])
                AOCLine.append(item_unitDict[key])
                AOCLine.append(quanitity)
                AOCLine.append(rate)
                def safeFloat(val, default = 0.0):
                    try:
                        return float(val)
                    except:
                        return default

                amount = float(safeFloat(quanitity )* safeFloat(rate))
                AOCLine.append(amount)
                AOCLine.append(None)
                AOCList.append(AOCLine)
            return AOCList
        def SumCCertiaIndex(aoc_list, index = -2):
            return sum(row[-2] for row in aoc_list if isinstance(row[-2], (int, float)))

        #___________________________________________________________________Calling and Writing
        row = Titles_writing(row)
        row+=1
        reduced_dict = datasplitting(segregated_dict)
        item_unitDict, item_rateDict = item_unit()
        AOCList = AOC_Data(reduced_dict)
        #Writing data to the Excel sheets
        for i, inner_row in enumerate(AOCList):
            for col, value in enumerate(inner_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.TNRnormalText_font

                # ✅ Alignment rules
                if col == 2:  # description column
                    cell.alignment = self.left_align
                else:
                    cell.alignment = self.center_align

                cell.border = self.thin_border
            row += 1  # move to next row after writing one line

        #___________________________________________________________________Processing and Writing
        def safe_float(val):
            try:
                return float(val)
            except:
                return val

        cont_, VAT_, physicalCont_, priceAdjCont_ = self.fetch_PrimaryKeys()[0]
        cont_, VAT_, physicalCont_, priceAdjCont_ = safe_float(cont_), safe_float(VAT_), safe_float(physicalCont_), safe_float(priceAdjCont_)
        print(cont_, VAT_, physicalCont_, priceAdjCont_ )


        TotalCivilAmount = SumCCertiaIndex(AOCList, index=-2)
        VATPercentage = VAT_/100 if  isinstance(VAT_, float) else 0.13
        VATAmount = VATPercentage* TotalCivilAmount

        #Contingencies amount ___________Compute after confirmation
        OfficeContigencyPercentage =  cont_/100 if  isinstance(cont_, float) else 0.13
        OfficeContigency = OfficeContigencyPercentage * TotalCivilAmount

        PhysicalContigencyPercentage =  physicalCont_/100 if  isinstance(physicalCont_, float) else 0.10
        PhysicalContigency = PhysicalContigencyPercentage * TotalCivilAmount

        PriceAdjContigencyPercentage =  priceAdjCont_/100 if  isinstance(priceAdjCont_, float) else 0.10
        PriceAdjContigency = PriceAdjContigencyPercentage * TotalCivilAmount


        TotelProjectCost = TotalCivilAmount + VATAmount + OfficeContigency + PhysicalContigency + PriceAdjContigency #Total project cost including, contigencies, PS, Vat and Civil works

        VerticalHeaders = ["Sub Total",
                           f"VAT Amount ({VATPercentage *100}% of Sub Total)",
                           f"Contingency ({OfficeContigencyPercentage * 100}% of Sub Total)",
                           f"Physical Contingency ({PhysicalContigencyPercentage * 100}% of Sub Total)",
                           f"Price Adjustment Contingency ({PriceAdjContigencyPercentage * 100}% of Sub Total)",
                           "Total Project Cost (including VAT PS & Contingencies)"]
        VerticalValues = [TotalCivilAmount, VATAmount, OfficeContigency, PhysicalContigency, PriceAdjContigency, TotelProjectCost]

        startrow =             row
        for header, value in zip(VerticalHeaders, VerticalValues):
            row = row
            ws.merge_cells(f"A{row}:E{row}")
            cellH = ws.cell(row=row, column=1, value=header) #Header column
            cellV = ws.cell(row=row, column=6, value=value) #Header column

            cellH.alignment = self.right_align
            cellH.font = self.TNRbold_font
            cellH.border = self.thin_border

            cellV.alignment = self.right_align
            cellV.font = self.TNRbold_font
            cellV.border = self.thin_border

            row+= 1

        columnsNO = len(AOCList[0])
        for Rowline in range(startrow, row):
            for col in range(1, columnsNO + 1):
                cell = ws.cell(row=Rowline, column=col)
                cell.border = self.bottom_border_thick

        wb.save(self.excel_name)
        print(f"Data written successfully to Abstract of Cost in {self.excel_name}")
        wb.close()
        return True

    def SummaryWriting(self):
        DescriptionTitle_index = 1  #Zero based indexing
        QuantityTitle_index = 7  #Zero based indexing
        Unit_Bulkindex = self.qEstDBtitles.index("unit")
        Rate_Bulkindex= self.qEstDBtitles.index("rate")
        wb = op.load_workbook(self.excel_name)  # 👈 change filename if needed
        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
        else:
            print("Sheet 'Summary' not found!")
            exit()
        row = ws.max_row + 1       #From here the writing needs to be started

        def Titles_writing(row):
            mergerrows = f"{self.sheetnamesTITLEMerger_Range['Summary'][0]}{row}:{self.sheetnamesTITLEMerger_Range['Summary'][1]}{row}"
            ws.merge_cells(mergerrows)
            cell = ws.cell(row=row, column=1)
            cell.value = f"Quantity Summary"
            cell.font = self.Special_font
            cell.alignment = self.center_align
            row += 1

            # Writing for the titles of estimation
            for col, header in enumerate(self.Summaryheaders, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.TNRbold_font
                cell.alignment = self.center_align
                cell.border = self.thin_border
                cell.border = self.bottom_border_thick


            # Optional: Adjust column widths for readability
            column_widths = [5, 38, 10, 10, 10]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width
            return row
        def datasplitting(segregated_dict):
            reduced_dict = {}
            for key, rows in segregated_dict.items():
                if rows:  # make sure not empty
                    reduced_dict[key] = [rows[0], rows[-1]]
            return reduced_dict
        def item_unit():
            item_unitDict = {}
            item_rateDict = {}
            for key, value in categorizedData.items():
                item_unitDict[key] = value[0][Unit_Bulkindex]
                item_rateDict[key] = value[0][Rate_Bulkindex]


            return item_unitDict, item_rateDict
        def SummaryData(reduced_dict):
            AOCList  = []

            for key, value in reduced_dict.items():
                AOCLine = []  ### "S.N.", "Description of Works", "Unit", "Quantity", "Rate", "Amount" , "Remarks"
                quanitity = value[1][QuantityTitle_index]
                rate = item_rateDict[key]

                AOCLine.append(key)
                AOCLine.append(value[0][DescriptionTitle_index])
                AOCLine.append(quanitity)
                AOCLine.append(item_unitDict[key])
                AOCLine.append(None)

                AOCList.append(AOCLine)
            return AOCList

        #___________________________________________________________________Calling and Writing
        itemNo_List, trimmedData, categorizedData, bulkedData, segregated_dict = self.ProcessedData_Extractor()
        row = Titles_writing(row)
        row+=1
        reduced_dict = datasplitting(segregated_dict)
        item_unitDict, item_rateDict = item_unit()
        SummaryList = SummaryData(reduced_dict)
        #Writing data to the Excel sheets
        for i, inner_row in enumerate(SummaryList):
            for col, value in enumerate(inner_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.TNRnormalText_font
                # ✅ Alignment rules
                if col == 2:  # description column
                    cell.alignment = self.left_align
                else:
                    cell.alignment = self.center_align

                if col == 3:
                    cell.font = self.TNRbold_font

                cell.border = self.thin_border
            row += 1  # move to next row after writing one line


        wb.save(self.excel_name)
        print(f"Data written successfully to Summary in {self.excel_name}")
        wb.close()
        return True

    def BOQ_Writing(self):
        DescriptionTitle_index = 1  #Zero based indexing
        QuantityTitle_index = 7  #Zero based indexing
        Unit_Bulkindex = self.qEstDBtitles.index("unit")
        Rate_Bulkindex= self.qEstDBtitles.index("rate")
        wb = op.load_workbook(self.excel_name)  # 👈 change filename if needed
        if "BOQ" in wb.sheetnames:
            ws = wb["BOQ"]
        else:
            print("Sheet 'Summary' not found!")
            exit()
        row = ws.max_row + 1       #From here the writing needs to be started

        def Titles_writing(row):
            mergerrows = f"{self.sheetnamesTITLEMerger_Range['BOQ'][0]}{row}:{self.sheetnamesTITLEMerger_Range['BOQ'][1]}{row}"
            ws.merge_cells(mergerrows)
            cell = ws.cell(row=row, column=1)
            cell.value = f"Bill of Quantity {12}"
            cell.font = self.Special_font
            cell.alignment = self.center_align
            row += 1

            # Writing for the titles of estimation
            for col, header in enumerate(self.BOQheaders, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = self.TNRbold_font
                cell.alignment = self.center_align
                cell.border = self.thin_border
                cell.border = self.bottom_border_thick

            ws.row_dimensions[row].height = 30  # Set height for this row



            # Optional: Adjust column widths for readability
            column_widths = [4, 15, 6, 7, 11, 20, 11]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = width
            return row
        def datasplitting(segregated_dict):
            reduced_dict = {}
            for key, rows in segregated_dict.items():
                if rows:  # make sure not empty
                    reduced_dict[key] = [rows[0], rows[-1]]
            return reduced_dict
        def item_unit():
            item_unitDict = {}
            item_rateDict = {}
            for key, value in categorizedData.items():
                item_unitDict[key] = value[0][Unit_Bulkindex]
                item_rateDict[key] = value[0][Rate_Bulkindex]


            return item_unitDict, item_rateDict
        def BOQData(reduced_dict):
            AOCList  = []

            for key, value in reduced_dict.items():
                AOCLine = []  ### "S.N.", "Description of Works", "Unit", "Quantity", "Rate", "Amount" , "Remarks"
                quanitity = value[1][QuantityTitle_index]
                rate = item_rateDict[key]

                AOCLine.append(key)
                AOCLine.append(value[0][DescriptionTitle_index])
                AOCLine.append(item_unitDict[key])
                AOCLine.append(quanitity)
                AOCLine.append(None)
                AOCLine.append(None)
                AOCLine.append(None)

                AOCList.append(AOCLine)
            return AOCList

        #___________________________________________________________________Calling and Writing
        itemNo_List, trimmedData, categorizedData, bulkedData, segregated_dict = self.ProcessedData_Extractor()
        row = Titles_writing(row)
        row+=1
        reduced_dict = datasplitting(segregated_dict)
        item_unitDict, item_rateDict = item_unit()
        BOQList = BOQData(reduced_dict)
        #Writing data to the Excel sheets
        for i, inner_row in enumerate(BOQList):
            for col, value in enumerate(inner_row, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = self.TNRnormalText_font
                # ✅ Alignment rules
                if col == 2:  # description column
                    cell.alignment = self.left_align
                else:
                    cell.alignment = self.center_align

                if col == 3:
                    cell.font = self.TNRbold_font

                cell.border = self.thin_border
            row += 1  # move to next row after writing one line

        #___________________________________________________________________ Processing and Writing
        LeftVerticalHeaders = ["Contractors Firm:", "Seal:", "Proprietor's Name:", "Address:", "Contact No.:", "Signature: ", "Date:"]
        RightVerticalHeaders = ["Total Amount (In figure):", "VAT (13%):", "Grand Total (In figure):", "Grand Total (In words):"]

        startrow = row
        row_heights = [25, 50, 25, 20, 20, 20, 20]
        columnnos = len(self.BOQheaders)
        for idx, (header, value) in enumerate(zip_longest(LeftVerticalHeaders, RightVerticalHeaders, fillvalue=""), start=1):
            row = row
            ws.merge_cells(f"A{row}:D{row}")
            ws.merge_cells(f"E{row}:G{row}")

            cellH = ws.cell(row=row, column=1, value=header) #Header column
            cellV = ws.cell(row=row, column=5, value=value) #Header column

            cellH.alignment = self.left_align
            cellH.font = self.TNRbold_font

            cellV.alignment = self.left_align
            cellV.font = self.TNRbold_font

            ws.row_dimensions[row].height = row_heights[idx-1]  # Set height for this row


            row +=1
        for row in range(startrow , row):
            for col in range(1, columnnos + 1):
                cell = ws.cell(row=row, column=col)

                cell.border = self.bottom_border_thick

        wb.save(self.excel_name)
        print(f"Data written successfully to BOQ in {self.excel_name}")
        wb.close()
        return True

    def RateAnalysisDataWriting(self):
        itemNo_List, trimmedData, categorizedData, bulkedData, segregated_dict = self.ProcessedData_Extractor()

        wb = op.load_workbook(self.excel_name)  # 👈 change filename if needed
        if "Rate Analysis" in wb.sheetnames:
            ws = wb["Rate Analysis"]
        else:
            print("Rate Analysis' not found!")
            exit()
        row = ws.max_row + 1       #From here the writing needs to be started
        datawritingStartRow = row

        RateAnalysisDict = {}
        for items in itemNo_List:
            rows, appliedRateData = self.db_GUI.load_appliedRateAnalysis(item_number=items)
            RateAnalysisDict[items] = {"rows": rows,
                                       "appliedRateData": appliedRateData}
            IntegratedData = appliedRateData

            def safeFloat(val, default=0.0):
                try:
                    return round(float(val), 2)
                except:
                    return default

            def RateAnalysisData(IntegratedData):
                # --- Write Top Section ---
                Row1 = [IntegratedData['NormsDBRef'], IntegratedData['Title_Section']['Title'][0]]
                Row2 = [IntegratedData['References']['Reference'][0], IntegratedData['Title_Section']['Note'][0]]
                TitlesRow = [Row1, Row2]

                # Header row
                Tabledyn1 = (
                        IntegratedData['First Inner table']['Title'] + IntegratedData['First Inner table']['Manpower'] +
                        IntegratedData['First Inner table']['Materials'] + IntegratedData['First Inner table'][
                            'Machines'] + IntegratedData['First Inner table']['Others'])
                Tablerows1 = []
                for row in Tabledyn1:
                    row.insert(0, None)
                    Tablerows1.append(row)

                # --- Totals Section ---
                columnnos = len(IntegratedData['First Inner table']['Title'][0])
                Row1 = [None] * columnnos
                Row1[columnnos - 3] = "वास्तविक दररेट"
                Row1[columnnos - 1] = IntegratedData['Second Inner table']['Original Rate'][0]

                Row2 = [None] * columnnos
                Row2[columnnos - 3] = "१५% ठेकदार ओभरहेड"
                Row2[columnnos - 1] = IntegratedData['Second Inner table']['Overhead'][0]
                Row2[1] = IntegratedData['Second Inner table']['Others'][0]

                Row3 = [None] * columnnos
                Row3[columnnos - 3] = "जम्मा दररेट"
                Row3[columnnos - 1] = IntegratedData['Second Inner table']['Total Rate'][0]
                Row3[0] = "रु."
                Row3[1] = safeFloat(IntegratedData['Second Inner table']['Unit Rate'][0])
                Row3[2] = "पै."

                TablesRow2 = [Row1, Row2, Row3]
                TablesRows = (Tablerows1 + TablesRow2)
                Tabledata = {"TitlesRow": TitlesRow, "TablesRows": TablesRows}

                return Tabledata

            def RateAnalsysisWriting(Tabledata, row):
                for index, line in enumerate(Tabledata["TitlesRow"]):
                    cell = ws.cell(row=row, column=1, value=line[0])
                    cell.font = self.TNRbold_font
                    cell.alignment = self.left_align
                    if index==0:
                        cell.border = self.borderA
                    else:
                        cell.border = self.borderC


                    ws.merge_cells(f"B{row}:H{row}")
                    cell = ws.cell(row=row, column=2, value=line[1])
                    cell.alignment = self.center_align
                    if index==0:
                        cell.font = self.KMbold_font
                        ws.row_dimensions[row].height = 40  # Set height for this row
                        for col in range(2, 9):  # B=2, H=8
                            cell = ws.cell(row=row, column=col, value=line[1] if col == 2 else None)
                            cell.border = self.borderB
                    else:
                        cell.font = self.KMNormal_font
                        for col in range(2, 9):  # B=2, H=8
                            cell = ws.cell(row=row, column=col, value=line[1] if col == 2 else None)
                            cell.border = self.borderD
                    row += 1
                def safeFloat(value):
                    try:
                        return round(float(value), 2)
                    except:
                        return value

                totallines = len(Tabledata['TablesRows'])
                for lineno, line in enumerate(Tabledata['TablesRows'], start=1):
                    for col, value in enumerate(line, start=1):
                        value = safeFloat(value)
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = self.KMNormal_font
                        cell.alignment = self.left_align

                        # ✅ Alignment rules
                        if lineno == 1:
                            cell.font = self.KMbold_font

                        if totallines - lineno < 3 :
                            cell.border = self.bottom_border_thick

                        else:
                            cell.border = self.thin_border

                        if totallines - lineno == 0:
                            if col == 2: #Rate cell
                                cell.border = self.thick_Redborder
                                cell.font = self.TNRbold_font
                                cell.alignment = self.center_align

                    row += 1  # move to next row after writing one line

                return row

            Tabledata = RateAnalysisData(IntegratedData)
            row = RateAnalsysisWriting(Tabledata, row)

            row+= 2

        #Data correction
        for r in range(datawritingStartRow, row + 1):
            cell = ws.cell(row=r, column=1)  # Column A
            cell.alignment = self.right_align

        column_widths = [4, 10, 12, 7, 6, 14, 8, 12]            #MAke total sum less than 73 else the text gets overflown
        for i, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = width

        wb.save(self.excel_name)
        print("Excel file created: RateAnalysis_Output.xlsx")
        return True

    def PostProcessingExcel(self):
        wb = op.load_workbook(self.excel_name)  # 👈 change filename if needed

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            row = ws.max_row

            #Set print area
            finalColumn = self.sheetnamesTITLEMerger_Range[sheet][1]
            freezingRow = self.sheetnamesTITLEMerger_Range[sheet][3]
            ws.print_area = f"A1:{finalColumn}{row}"

            # Optionally, you can also set page setup options
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.scale = 100  # ensure Excel applies scaling

            # Reduce margins
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.5
            ws.page_margins.bottom = 0.75

            ws.freeze_panes = ws[f"A{freezingRow}"]  # everything above row 8 is frozen (rows 1–7)
            ws.print_title_rows = f"1:{freezingRow-1}"  # rows 1-7 repeat on every printed page

            # --- Set footer to show page numbers ---
            ws.page_margins.footer = 0.05 # in inches, default is 0.3
            # &P = current page number, &N = total pages #&10 for the font size
            ws.oddFooter.left.text = "&10Prepared by\n"
            ws.oddFooter.center.text = "&10Checked by\nPage &P of &N"
            ws.oddFooter.right.text = "&10Approved by\n"



        wb.save(self.excel_name)
        print("Excel file formatted successfully")
        return True


    def excel_to_pdf_merge(self):

        app = xw.App(visible=False)
        wb = app.books.open(self.excel_name)
        temp_pdfs = []

        for sheet in wb.sheets:
            # ✅ enforce absolute path
            pdf_file = os.path.join(self.base_dir, f"{sheet.name}.pdf")

            # Export each sheet as PDF
            sheet.api.ExportAsFixedFormat(0, pdf_file)
            temp_pdfs.append(pdf_file)

        wb.close()
        app.quit()


        # Merge PDFs
        merger = PdfMerger()
        for pdf in temp_pdfs:
            merger.append(pdf)
        merger.write(self.pdf_path)
        merger.close()

        # Cleanup
        for pdf in temp_pdfs:
            os.remove(pdf)
        return True

    def PrintPDF(self):
        os.startfile(self.pdf_path, "print")         #Ensure to set the default printer from settings(set windows chooes = false, and set the printer as default)
        return True


# db_out = DB_Output()
# db_out.SheetsTitle_Writing()    #When the erroro of the io ie file not found occurs then enforce the Sheets Title function
# db_out.CoverPage_Writing()
# db_out.QuantityEstSheet_Writing()
# db_out.AOC_writing()
# db_out.SummaryWriting()
# db_out.BOQ_Writing()
# db_out.RateAnalysisDataWriting()
# db_out.PostProcessingExcel()
# db_out.excel_to_pdf_merge()






